# -*- coding: utf-8 -*-
"""
Script para cargar catalogos del RESP desde el archivo Excel.

USO:
    python cargar_catalogos.py
    python cargar_catalogos.py "C:\\ruta\\al\\Catalogos_Sistema_RESP.xlsx"

Coloca el archivo Excel en cualquiera de estas ubicaciones y se detectara
automaticamente:
    - Misma carpeta que este script
    - Carpeta padre
    - Escritorio de Windows
    - Carpeta Descargas de Windows
"""
import os
import sys
import django

# ── Configuracion Django ──────────────────────────────────────────────────────
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'resp_project.settings')
# Limpiar variables PG de Windows para evitar UnicodeDecodeError
for _v in ('PGUSER', 'PGPASSWORD', 'PGHOST', 'PGDATABASE', 'PGPORT'):
    os.environ.pop(_v, None)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)
django.setup()

# ── Localizacion del Excel ────────────────────────────────────────────────────
NOMBRE_EXCEL = 'Catalogos_Sistema_RESP.xlsx'

def encontrar_excel():
    # 1. Argumento de linea de comandos
    if len(sys.argv) > 1:
        ruta = sys.argv[1].strip('"').strip("'")
        if os.path.isfile(ruta):
            return ruta
        else:
            print(f"ERROR: No se encontro el archivo en: {ruta}")
            sys.exit(1)

    # 2. Misma carpeta que el script
    candidatos = [
        os.path.join(BASE_DIR, NOMBRE_EXCEL),
        os.path.join(BASE_DIR, '..', NOMBRE_EXCEL),
    ]

    # 3. Escritorio y Descargas de Windows
    userprofile = os.environ.get('USERPROFILE', '')
    if userprofile:
        candidatos += [
            os.path.join(userprofile, 'Desktop', NOMBRE_EXCEL),
            os.path.join(userprofile, 'Escritorio', NOMBRE_EXCEL),
            os.path.join(userprofile, 'Downloads', NOMBRE_EXCEL),
            os.path.join(userprofile, 'Descargas', NOMBRE_EXCEL),
        ]

    for c in candidatos:
        if os.path.isfile(c):
            return os.path.abspath(c)

    # No encontrado - pedir ruta al usuario
    print("\n" + "="*60)
    print("  No se encontro: " + NOMBRE_EXCEL)
    print("="*60)
    print("\nOpciones:")
    print("  1. Copia el archivo Excel a la misma carpeta que este script")
    print("  2. Ejecuta: python cargar_catalogos.py \"ruta\\al\\archivo.xlsx\"")
    print("\nEjemplo:")
    print('  python cargar_catalogos.py "C:\\Users\\TuUsuario\\Downloads\\Catalogos_Sistema_RESP.xlsx"')
    print()
    # Intentar pedir ruta interactiva
    try:
        ruta = input("O ingresa la ruta ahora: ").strip().strip('"').strip("'")
        if os.path.isfile(ruta):
            return ruta
    except (EOFError, KeyboardInterrupt):
        pass
    print("No se pudo localizar el archivo. Abortando.")
    sys.exit(1)


EXCEL = encontrar_excel()
print(f"\nUsando archivo: {EXCEL}\n")

# ── Carga de catalogos ────────────────────────────────────────────────────────
from openpyxl import load_workbook
from catalogos.models import (
    FuenteFinanciamiento, Dependencia, UnidadAdministrativa,
    Programa, Proyecto,
    TipoContratacion, TipoPersonal, TipoFuncion, NivelEstructura,
    EstatusPlaza, TipoDeclaracion, Area, NivelEscolaridad,
    Discapacidad, EnfermedadCronica, Pueblo, MotivoBaja, Idioma,
    EstadoCivil, Pais, EntidadFederativa, Municipio, Sindicato, Categoria
)

wb = load_workbook(EXCEL, read_only=True, data_only=True)
print(f"Hojas encontradas: {wb.sheetnames}\n")


def get_rows(sheet_name, skip=2):
    if sheet_name not in wb.sheetnames:
        print(f"  [WARN] Hoja no encontrada: {sheet_name}")
        return []
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip:
            continue
        if any(v is not None for v in row):
            rows.append(row)
    return rows


def safe_str(v, default=''):
    if v is None:
        return default
    return str(v).strip()


def safe_int(v, default=0):
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


resultados = {}

# 1. Fuentes de Financiamiento
print("Cargando Fuentes de Financiamiento...")
c = 0
for row in get_rows('Fuente Financiamiento'):
    clave, desc = row[0], row[1]
    if clave and desc:
        _, created = FuenteFinanciamiento.objects.get_or_create(
            clave=safe_str(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Fuentes de Financiamiento'] = FuenteFinanciamiento.objects.count()
print(f"  {c} nuevos / {resultados['Fuentes de Financiamiento']} total")

# 2. Dependencias
print("Cargando Dependencias...")
c = 0
for row in get_rows('Dependencias'):
    ejercicio, clave, desc = row[0], row[1], row[2]
    if clave and desc:
        _, created = Dependencia.objects.get_or_create(
            ejercicio=safe_int(ejercicio, 2025),
            clave=safe_str(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Dependencias'] = Dependencia.objects.count()
print(f"  {c} nuevas / {resultados['Dependencias']} total")

# 3. Categorias
print("Cargando Categorias...")
c = 0
for row in get_rows('Categoria'):
    row_ext = (row + (None,) * 7)[:7]
    _, clave, subca, desc, tipo_plaza, tp, nivel = row_ext
    if clave and desc:
        _, created = Categoria.objects.get_or_create(
            clave=safe_str(clave),
            defaults={
                'descripcion': safe_str(desc),
                'subcategoria': safe_int(subca, 0),
                'tipo_plaza': safe_str(tipo_plaza),
                'tp': safe_str(tp),
                'nivel': safe_int(nivel, 0),
            }
        )
        if created:
            c += 1
resultados['Categorias'] = Categoria.objects.count()
print(f"  {c} nuevas / {resultados['Categorias']} total")

# 4. Tipos de Contratacion
print("Cargando Tipos de Contratacion...")
c = 0
# La hoja puede llamarse 'Tipo Contratacion' o 'Tipo Contratacion' (con acento)
_sheet_tc = 'Tipo Contratación' if 'Tipo Contratación' in wb.sheetnames else 'Tipo Contratacion'
for row in get_rows(_sheet_tc):
    if not row or row[0] is None:
        continue
    clave, desc = row[0], row[1]
    if clave and desc:
        _, created = TipoContratacion.objects.get_or_create(
            clave=safe_str(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Tipos Contratacion'] = TipoContratacion.objects.count()
print(f"  {c} nuevos / {resultados['Tipos Contratacion']} total")

# 5. Tipos de Funcion
print("Cargando Tipos de Funcion...")
c = 0
_sheet_tf = 'Tipo Función' if 'Tipo Función' in wb.sheetnames else 'Tipo Funcion'
for row in get_rows(_sheet_tf):
    if not row:
        continue
    clave = row[0]
    desc = row[2] if len(row) > 2 else row[1]
    if clave and desc:
        _, created = TipoFuncion.objects.get_or_create(
            clave=safe_str(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Tipos Funcion'] = TipoFuncion.objects.count()
print(f"  {c} nuevos / {resultados['Tipos Funcion']} total")

# 6. Niveles de Estructura
print("Cargando Niveles de Estructura...")
c = 0
for row in get_rows('Nivel Estructura'):
    row_ext = (row + (None,) * 3)[:3]
    clave, desc, nivel = row_ext
    if clave and desc:
        _, created = NivelEstructura.objects.get_or_create(
            clave=safe_str(clave),
            defaults={
                'descripcion': safe_str(desc),
                'nivel': safe_int(nivel, 0),
            }
        )
        if created:
            c += 1
resultados['Niveles Estructura'] = NivelEstructura.objects.count()
print(f"  {c} nuevos / {resultados['Niveles Estructura']} total")

# 7. Estatus de Plaza
print("Cargando Estatus de Plaza...")
c = 0
for row in get_rows('estatus plaza'):
    clave, desc = row[0], row[1]
    if clave and desc:
        _, created = EstatusPlaza.objects.get_or_create(
            clave=safe_str(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Estatus Plaza'] = EstatusPlaza.objects.count()
print(f"  {c} nuevos / {resultados['Estatus Plaza']} total")

# 8. Tipos de Declaracion
print("Cargando Tipos de Declaracion...")
c = 0
for row in get_rows('Tipo_Declaracion'):
    clave, desc = row[0], row[1]
    if clave and desc:
        _, created = TipoDeclaracion.objects.get_or_create(
            clave=safe_str(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Tipos Declaracion'] = TipoDeclaracion.objects.count()
print(f"  {c} nuevos / {resultados['Tipos Declaracion']} total")

# 9. Areas
print("Cargando Areas...")
c = 0
for row in get_rows('Areas'):
    clave, desc = row[0], row[1]
    if clave and desc:
        _, created = Area.objects.get_or_create(
            clave=safe_str(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Areas'] = Area.objects.count()
print(f"  {c} nuevas / {resultados['Areas']} total")

# 10. Niveles de Escolaridad
print("Cargando Niveles de Escolaridad...")
c = 0
for row in get_rows('Nivel_Escolaridad'):
    clave, desc, estatus = (row + (None,) * 3)[:3]
    if clave is not None and desc:
        _, created = NivelEscolaridad.objects.get_or_create(
            clave=safe_int(clave),
            defaults={
                'descripcion': safe_str(desc),
                'estatus': safe_str(estatus),
            }
        )
        if created:
            c += 1
resultados['Niveles Escolaridad'] = NivelEscolaridad.objects.count()
print(f"  {c} nuevos / {resultados['Niveles Escolaridad']} total")

# 11. Discapacidades
print("Cargando Discapacidades...")
c = 0
for row in get_rows('Discapacidad'):
    row_ext = (row + (None,) * 4)[:4]
    clave, tipo, desc, _ = row_ext
    if clave is not None and tipo:
        _, created = Discapacidad.objects.get_or_create(
            clave=safe_int(clave),
            defaults={
                'tipo': safe_str(tipo),
                'descripcion': safe_str(desc),
            }
        )
        if created:
            c += 1
resultados['Discapacidades'] = Discapacidad.objects.count()
print(f"  {c} nuevas / {resultados['Discapacidades']} total")

# 12. Enfermedades Cronicas
print("Cargando Enfermedades Cronicas...")
c = 0
for row in get_rows('Enferm Cronicas'):
    clave, desc = row[0], row[1]
    if clave and desc:
        _, created = EnfermedadCronica.objects.get_or_create(
            clave=safe_int(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Enfermedades Cronicas'] = EnfermedadCronica.objects.count()
print(f"  {c} nuevas / {resultados['Enfermedades Cronicas']} total")

# 13. Pueblos Indigenas
print("Cargando Pueblos Indigenas/Afromexicanos...")
c = 0
for row in get_rows('Pueblos'):
    clave, desc = row[0], row[1]
    if clave and desc:
        _, created = Pueblo.objects.get_or_create(
            clave=safe_int(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Pueblos'] = Pueblo.objects.count()
print(f"  {c} nuevos / {resultados['Pueblos']} total")

# 14. Motivos de Baja
print("Cargando Motivos de Baja...")
c = 0
for row in get_rows('Motivos Baja'):
    clave, desc = row[0], row[1]
    if clave and desc:
        _, created = MotivoBaja.objects.get_or_create(
            clave=safe_str(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Motivos Baja'] = MotivoBaja.objects.count()
print(f"  {c} nuevos / {resultados['Motivos Baja']} total")

# 15. Idiomas
print("Cargando Idiomas/Lenguas...")
c = 0
for row in get_rows('Idiomas'):
    row_ext = (row + (None,) * 4)[:4]
    clave, id_cndh, desc, familia = row_ext
    if clave and desc:
        _, created = Idioma.objects.get_or_create(
            clave=safe_int(clave),
            defaults={
                'identificador_cndh': safe_int(id_cndh) if id_cndh else None,
                'descripcion': safe_str(desc),
                'familia_linguistica': safe_str(familia),
            }
        )
        if created:
            c += 1
resultados['Idiomas'] = Idioma.objects.count()
print(f"  {c} nuevos / {resultados['Idiomas']} total")

# 16. Estado Civil
print("Cargando Estados Civiles...")
c = 0
for row in get_rows('estado civil'):
    clave, desc = row[0], row[1]
    if clave and desc:
        _, created = EstadoCivil.objects.get_or_create(
            clave=safe_str(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Estados Civiles'] = EstadoCivil.objects.count()
print(f"  {c} nuevos / {resultados['Estados Civiles']} total")

# 17. Paises
print("Cargando Paises...")
c = 0
for row in get_rows('paises'):
    clave, nombre = row[0], row[1]
    if clave and nombre:
        _, created = Pais.objects.get_or_create(
            clave=safe_int(clave),
            defaults={'nombre': safe_str(nombre)}
        )
        if created:
            c += 1
resultados['Paises'] = Pais.objects.count()
print(f"  {c} nuevos / {resultados['Paises']} total")

# 18. Entidades Federativas
print("Cargando Entidades Federativas...")
c = 0
for row in get_rows('Entidades Federativas'):
    row_ext = (row + (None,) * 3)[:3]
    clave, abrev, nombre = row_ext
    if clave and nombre:
        _, created = EntidadFederativa.objects.get_or_create(
            clave=safe_int(clave),
            defaults={
                'abreviatura': safe_str(abrev),
                'nombre': safe_str(nombre),
            }
        )
        if created:
            c += 1
resultados['Entidades Federativas'] = EntidadFederativa.objects.count()
print(f"  {c} nuevas / {resultados['Entidades Federativas']} total")

# 19. Municipios
print("Cargando Municipios (puede tardar unos segundos)...")
c = 0
errores_mun = 0
for row in get_rows('Municipios'):
    row_ext = (row + (None,) * 4)[:4]
    abrev, ent_clave, mun_clave, nombre = row_ext
    if ent_clave and mun_clave and nombre:
        try:
            ent = EntidadFederativa.objects.get(clave=safe_int(ent_clave))
            _, created = Municipio.objects.get_or_create(
                entidad=ent,
                clave=safe_str(mun_clave),
                defaults={
                    'abreviatura': safe_str(abrev),
                    'nombre': safe_str(nombre),
                }
            )
            if created:
                c += 1
        except EntidadFederativa.DoesNotExist:
            errores_mun += 1
resultados['Municipios'] = Municipio.objects.count()
print(f"  {c} nuevos / {resultados['Municipios']} total  ({errores_mun} entidades no encontradas)")

# 20. Sindicatos
print("Cargando Sindicatos...")
c = 0
for row in get_rows('Sindicatos'):
    row_ext = (row + (None,) * 3)[:3]
    _, clave, desc = row_ext
    if clave and desc:
        _, created = Sindicato.objects.get_or_create(
            clave=safe_str(clave),
            defaults={'descripcion': safe_str(desc)}
        )
        if created:
            c += 1
resultados['Sindicatos'] = Sindicato.objects.count()
print(f"  {c} nuevos / {resultados['Sindicatos']} total")


# 21. Unidades Administrativas
print("Cargando Unidades Administrativas...")
c = 0
for row in get_rows('Unidades Admvas'):
    # Cols: EJERCICIO, Depcia, Clave_Unidad, Descripcion_unidad
    row_ext = (row + (None,) * 4)[:4]
    ejercicio, dep_clave, clave_uni, desc = row_ext
    if not clave_uni or not desc:
        continue
    dep_clave = safe_str(dep_clave)
    dep = Dependencia.objects.filter(clave=dep_clave).first()
    if not dep:
        continue
    _, created = UnidadAdministrativa.objects.get_or_create(
        clave=safe_str(clave_uni),
        defaults={
            'ejercicio':   safe_int(ejercicio, 2025),
            'dependencia': dep,
            'descripcion': safe_str(desc),
        }
    )
    if created:
        c += 1
resultados['Unidades Administrativas'] = UnidadAdministrativa.objects.count()
print(f"  {c} nuevas / {resultados['Unidades Administrativas']} total")


# ── Proyecto por defecto (clave 00000000) ────────────────────────────────────
# Cuando el layout trae PROYECTO = 00000000 significa "sin proyecto específico".
# Se crea un Programa y Proyecto comodín por cada Dependencia registrada.
print("Creando Programa/Proyecto por defecto (00000000)...")
_prg_default_desc = "PROYECTO GENERAL"
_pry_default_clave = "00000000"
_dep_default_count = 0

for _dep in Dependencia.objects.all():
    # Buscar o crear una unidad comodín para esta dependencia si no existe ninguna
    _uni = UnidadAdministrativa.objects.filter(dependencia=_dep).first()
    if not _uni:
        continue   # sin unidad no podemos crear el programa

    # Programa default
    _prg, _ = Programa.objects.get_or_create(
        clave="DEFAULT",
        unidad=_uni,
        defaults={
            'ejercicio':   2025,
            'dependencia': _dep,
            'descripcion': "PROGRAMA GENERAL",
        }
    )

    # Proyecto default 00000000
    _, created = Proyecto.objects.get_or_create(
        clave=_pry_default_clave,
        programa=_prg,
        defaults={
            'ejercicio':   2025,
            'dependencia': _dep,
            'unidad':      _uni,
            'descripcion': _prg_default_desc,
        }
    )
    if created:
        _dep_default_count += 1

print(f"  {_dep_default_count} proyectos default creados")
# ─────────────────────────────────────────────────────────────────────────────

# 22. Programas
print("Cargando Programas...")
c = 0
errores_prg = 0
for row in get_rows('programas'):
    # Cols: EJERCICIO, ID_DEPCIA, ID_UNIDAD, CLAVE_PROGRAMA, DESCRIPCION
    row_ext = (row + (None,) * 5)[:5]
    ejercicio, dep_clave, uni_clave, clave_prg, desc = row_ext
    if not clave_prg or not desc:
        continue
    dep = Dependencia.objects.filter(clave=safe_str(dep_clave)).first()
    uni = UnidadAdministrativa.objects.filter(clave=safe_str(uni_clave)).first()
    if not dep or not uni:
        errores_prg += 1
        continue
    _, created = Programa.objects.get_or_create(
        clave=safe_str(clave_prg),
        unidad=uni,
        defaults={
            'ejercicio':   safe_int(ejercicio, 2025),
            'dependencia': dep,
            'descripcion': safe_str(desc),
        }
    )
    if created:
        c += 1
resultados['Programas'] = Programa.objects.count()
print(f"  {c} nuevos / {resultados['Programas']} total  ({errores_prg} dependencia/unidad no encontrada)")

# 23. Proyectos
print("Cargando Proyectos...")
c = 0
errores_pry = 0
for row in get_rows('proyecto'):
    # Cols: LLAVE, EJERCICIO, ID_DEPCIA, ID_UNIDAD, ID_PROGRAMA, CLAVE_PROYECTO, DESCRIPCION
    row_ext = (row + (None,) * 7)[:7]
    _, ejercicio, dep_clave, uni_clave, prg_clave, clave_pry, desc = row_ext
    if not clave_pry or not desc:
        continue
    dep = Dependencia.objects.filter(clave=safe_str(dep_clave)).first()
    uni = UnidadAdministrativa.objects.filter(clave=safe_str(uni_clave)).first()
    prg = Programa.objects.filter(clave=safe_str(prg_clave), unidad=uni).first() if uni else None
    if not dep or not uni or not prg:
        errores_pry += 1
        continue
    _, created = Proyecto.objects.get_or_create(
        clave=safe_str(clave_pry).strip(),
        programa=prg,
        defaults={
            'ejercicio':   safe_int(ejercicio, 2025),
            'dependencia': dep,
            'unidad':      uni,
            'descripcion': safe_str(desc),
        }
    )
    if created:
        c += 1
resultados['Proyectos'] = Proyecto.objects.count()
print(f"  {c} nuevos / {resultados['Proyectos']} total  ({errores_pry} dependencia/unidad/programa no encontrado)")


# ── Resumen final ─────────────────────────────────────────────────────────────
print("\n" + "="*60)
print("  CATALOGOS CARGADOS EXITOSAMENTE")
print("="*60)
for nombre, total in resultados.items():
    print(f"  {nombre:<30} {total:>6} registros")
print("="*60)
print("\nSiguiente paso:")
print("  python manage.py setup_resp   <- crea usuario administrador")
print("  python manage.py runserver    <- inicia el servidor")
