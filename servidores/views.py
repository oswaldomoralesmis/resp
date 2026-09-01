# -*- coding: utf-8 -*-
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.urls import reverse_lazy
from django.db import transaction, connection
from django.core.management.color import no_style
from django.db.models import Q, Count, Value
from django.db.models.functions import Concat
from django.contrib import messages
from django.http import Http404, HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .models import (
    ServidorPublico, InformacionBasica, BajaServidorPublico, Puesto,
    DatosPersonales, DatosComplementarios, DiscapacidadServidor,
    EnfermedadCronicaServidor, IdiomaServidor, LogEvento,
    sincronizar_puesto, liberar_puestos_de,
)
from .forms import (
    ServidorPublicoForm, DatosPersonalesForm, DatosComplementariosForm,
    InformacionBasicaForm, BajaForm, PuestoForm,
)
from catalogos.models import Dependencia, EstatusPlaza, Discapacidad, EnfermedadCronica, Idioma
from cargas.models import PeriodoCarga, CargaLayout, periodo_vigente_hoy
from usuarios.mixins import DependenciaScopedMixin, filtrar_por_dependencia, admin_requerido, PermisoRequeridoMixin


@login_required
def redirect_to_dashboard(request):
    return redirect('dashboard')

@login_required
def dashboard(request):
    total_servidores = ServidorPublico.objects.filter(activo=True).count()
    total_dependencias = Dependencia.objects.count()
    # últimas cargas
    periodo_actual = periodo_vigente_hoy()
    # estadísticas por estatus
    stats_estatus = InformacionBasica.objects.filter(
        activo=True
    ).values('estatus_plaza__descripcion').annotate(total=Count('id')).order_by('-total')[:5]

    context = {
        'total_servidores': total_servidores,
        'total_dependencias': total_dependencias,
        'periodo_actual': periodo_actual,
        'stats_estatus': stats_estatus,
        'titulo': 'Dashboard',
        'debug_mode': settings.DEBUG,
    }
    return render(request, 'dashboard.html', context)


# (clave, etiqueta, modelo, claves de las que depende por integridad
# referencial PROTECT — deben eliminarse junto con ella). El orden de la
# lista ya respeta esas dependencias, así que basta con recorrerla en orden.
GRUPOS_RESET = [
    ('informacion_basica', 'Información Básica (historial quincenal)', InformacionBasica, []),
    ('bajas', 'Bajas de Servidores', BajaServidorPublico, []),
    ('plazas', 'Plazas', Puesto, []),
    ('servidores', 'Servidores Públicos', ServidorPublico, ['informacion_basica', 'bajas']),
    ('cargas', 'Cargas de Layouts', CargaLayout, []),
    ('periodos', 'Períodos de Carga', PeriodoCarga, ['cargas']),
]


@login_required
@admin_requerido
def reset_datos_prueba(request):
    """Borra, grupo por grupo a elección del administrador, los datos
    transaccionales (información básica, bajas, plazas, servidores, cargas y
    períodos) para dejar el ambiente listo para volver a probar desde cero.
    NO toca catálogos ni usuarios. Solo disponible con DEBUG=True (ambiente
    de desarrollo/pruebas), nunca en producción."""
    if not settings.DEBUG:
        raise Http404()

    grupos = [
        {'clave': clave, 'etiqueta': etiqueta, 'cantidad': modelo.objects.count()}
        for clave, etiqueta, modelo, _requiere in GRUPOS_RESET
    ]

    if request.method == 'POST':
        seleccionados = {
            clave for clave, _e, _m, _r in GRUPOS_RESET
            if request.POST.get(f'grupo_{clave}') == 'on'
        }
        if not seleccionados:
            messages.error(request, 'Seleccione al menos un grupo de datos a eliminar.')
            return redirect('reset_datos_prueba')

        # Un grupo con dependientes (p.ej. Servidores Públicos protege su
        # Información Básica y sus Bajas) solo se puede eliminar si esos
        # dependientes también están seleccionados; si no, la eliminación
        # fallaría a mitad de camino por la restricción PROTECT de la BD.
        etiquetas = {clave: etiqueta for clave, etiqueta, _m, _r in GRUPOS_RESET}
        faltantes = {}
        for clave in seleccionados:
            requiere = next(r for c, _e, _m, r in GRUPOS_RESET if c == clave)
            faltan = [etiquetas[r] for r in requiere if r not in seleccionados]
            if faltan:
                faltantes[etiquetas[clave]] = faltan
        if faltantes:
            detalle = '; '.join(
                f'"{grupo}" requiere también: {", ".join(reqs)}'
                for grupo, reqs in faltantes.items()
            )
            messages.error(
                request,
                f'No se puede eliminar por separado, por integridad de datos. {detalle}.'
            )
            return redirect('reset_datos_prueba')

        reiniciar_ids = request.POST.get('reiniciar_ids') == 'on'
        modelos_eliminados = [m for c, _e, m, _r in GRUPOS_RESET if c in seleccionados]
        with transaction.atomic():
            for modelo in modelos_eliminados:
                modelo.objects.all().delete()
            if reiniciar_ids:
                with connection.cursor() as cursor:
                    for sql in connection.ops.sequence_reset_sql(no_style(), modelos_eliminados):
                        cursor.execute(sql)

        nombres = ', '.join(etiquetas[c] for c in seleccionados)
        if reiniciar_ids:
            mensaje = (
                f'Se eliminó: {nombres}. Se reiniciaron sus IDs: los próximos registros que se '
                'generen comenzarán de nuevo en 1. Los catálogos y los usuarios del sistema no '
                'se modificaron.'
            )
        else:
            mensaje = (
                f'Se eliminó: {nombres}. Los catálogos y los usuarios del sistema no se modificaron.'
            )
        messages.success(request, mensaje)
        return redirect('dashboard')

    return render(request, 'servidores/reset_confirm.html', {
        'titulo': 'Reiniciar Datos de Prueba',
        'grupos': grupos,
        'total': sum(g['cantidad'] for g in grupos),
    })


class ServidorListView(LoginRequiredMixin, PermisoRequeridoMixin, ListView):
    permiso_clave = 'padron'
    model = ServidorPublico
    template_name = 'servidores/list.html'
    context_object_name = 'servidores'
    paginate_by = 20

    def get_queryset(self):
        qs = ServidorPublico.objects.all()
        estatus = self.request.GET.get('estatus', 'activos')
        if estatus == 'activos':
            qs = qs.filter(activo=True)
        elif estatus == 'inactivos':
            qs = qs.filter(activo=False)
        # 'todos': sin filtro de activo.

        # Ojo: el filtro de dependencia NO exige informacion_basica__activo=True
        # — un servidor dado de baja tiene su Información Básica desactivada,
        # pero sigue perteneciendo históricamente a esa dependencia. Exigir
        # activo=True aquí lo dejaría invisible incluso para su propia
        # dependencia al buscarlo como inactivo/baja.
        user = self.request.user
        if not user.es_administrador:
            qs = qs.filter(informacion_basica__dependencia_id=user.dependencia_id)
        q = self.request.GET.get('q', '')
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q) | Q(primer_apellido__icontains=q) |
                Q(rfc__icontains=q) | Q(curp__icontains=q) | Q(expediente__icontains=q)
            )
        dep = self.request.GET.get('dependencia', '')
        if dep:
            qs = qs.filter(informacion_basica__dependencia_id=dep)
        return qs.distinct()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['dependencias'] = filtrar_por_dependencia(Dependencia.objects.all(), self.request.user, 'pk')
        ctx['q'] = self.request.GET.get('q', '')
        ctx['estatus'] = self.request.GET.get('estatus', 'activos')
        ctx['titulo'] = 'Padrón de Servidores Públicos'
        return ctx


class ServidorDetailView(LoginRequiredMixin, DependenciaScopedMixin, DetailView):
    model = ServidorPublico
    template_name = 'servidores/detail.html'
    context_object_name = 'servidor'
    dependencia_lookup = 'informacion_basica__dependencia_id'

    def get_queryset(self):
        return super().get_queryset().distinct()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['info_basica'] = InformacionBasica.objects.filter(
            servidor=self.object, activo=True
        ).order_by('-quincena').first()
        ctx['historial'] = InformacionBasica.objects.filter(
            servidor=self.object
        ).order_by('-quincena')[:10]
        ctx['bajas'] = self.object.bajas.select_related('motivo_baja', 'dependencia').order_by('-fecha_baja')
        ctx['eventos_log'] = self.object.eventos_log.select_related('usuario', 'carga')[:30]
        ctx['titulo'] = f'Servidor: {self.object.nombre_completo}'
        return ctx


class ServidorCreateView(LoginRequiredMixin, CreateView):
    model = ServidorPublico
    form_class = ServidorPublicoForm
    template_name = 'servidores/form.html'
    success_url = reverse_lazy('servidor_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Alta de Servidor Público'
        ctx['accion'] = 'Registrar'
        return ctx


class ServidorUpdateView(LoginRequiredMixin, DependenciaScopedMixin, UpdateView):
    """Edición de un Servidor Público, incluyendo los datos de la sección
    'Datos del Servidor Público' (domicilio, escolaridad, discapacidades,
    pueblo indígena, enfermedades crónicas e idiomas), que solo tiene sentido
    capturar sobre un servidor ya existente."""
    model = ServidorPublico
    form_class = ServidorPublicoForm
    template_name = 'servidores/form.html'
    dependencia_lookup = 'informacion_basica__dependencia_id'

    def get_queryset(self):
        return super().get_queryset().distinct()

    def get_success_url(self):
        return reverse_lazy('servidor_detail', kwargs={'pk': self.object.pk})

    @staticmethod
    def _checkboxes(catalogo_qs, seleccionados, niveles=None):
        """Lista de {pk, texto, checked, nivel} para renderizar un checklist.
        Normaliza a texto porque 'seleccionados'/'niveles' vienen a veces de
        POST (strings) y a veces de la BD (ids)."""
        seleccionados = {str(s) for s in seleccionados}
        niveles = {str(k): v for k, v in (niveles or {}).items()}
        return [
            {
                'pk': item.pk,
                'texto': str(item),
                'checked': str(item.pk) in seleccionados,
                'nivel': niveles.get(str(item.pk), ''),
            }
            for item in catalogo_qs
        ]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f'Modificar: {self.object.nombre_completo}'
        ctx['accion'] = 'Guardar cambios'
        ctx.setdefault('form_personales', DatosPersonalesForm(instance=getattr(self.object, 'datos_personales', None)))
        ctx.setdefault('form_complementarios', DatosComplementariosForm(instance=getattr(self.object, 'datos_complementarios', None)))
        ctx.setdefault('discapacidades', self._checkboxes(
            Discapacidad.objects.all(), set(self.object.discapacidades.values_list('discapacidad_id', flat=True))
        ))
        ctx.setdefault('enfermedades', self._checkboxes(
            EnfermedadCronica.objects.all(), set(self.object.enfermedades.values_list('enfermedad_id', flat=True))
        ))
        ctx.setdefault('idiomas', self._checkboxes(
            Idioma.objects.all(), set(self.object.idiomas.values_list('idioma_id', flat=True)),
            niveles={r.idioma_id: r.nivel for r in self.object.idiomas.all()}
        ))
        return ctx

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        form_personales = DatosPersonalesForm(request.POST, instance=getattr(self.object, 'datos_personales', None))
        form_complementarios = DatosComplementariosForm(request.POST, instance=getattr(self.object, 'datos_complementarios', None))

        if form.is_valid() and form_personales.is_valid() and form_complementarios.is_valid():
            self.object = form.save()

            datos_personales_es_nuevo = form_personales.instance.pk is None
            datos_personales = form_personales.save(commit=False)
            datos_personales.servidor = self.object
            datos_personales.save()
            LogEvento.registrar(
                'datos_personales', datos_personales.pk, self.object,
                'creado' if datos_personales_es_nuevo else 'editado', 'manual', usuario=request.user,
            )

            datos_complementarios = form_complementarios.save(commit=False)
            datos_complementarios.servidor = self.object
            datos_complementarios.save()

            disc_ids = request.POST.getlist('discapacidades')
            self.object.discapacidades.all().delete()
            DiscapacidadServidor.objects.bulk_create([
                DiscapacidadServidor(servidor=self.object, discapacidad_id=pk) for pk in disc_ids
            ])

            enf_ids = request.POST.getlist('enfermedades')
            self.object.enfermedades.all().delete()
            EnfermedadCronicaServidor.objects.bulk_create([
                EnfermedadCronicaServidor(servidor=self.object, enfermedad_id=pk) for pk in enf_ids
            ])

            idioma_ids = request.POST.getlist('idiomas')
            self.object.idiomas.all().delete()
            IdiomaServidor.objects.bulk_create([
                IdiomaServidor(servidor=self.object, idioma_id=pk, nivel=request.POST.get(f'nivel_idioma_{pk}', '').strip())
                for pk in idioma_ids
            ])

            return redirect(self.get_success_url())

        niveles_post = {
            pk: request.POST.get(f'nivel_idioma_{pk}', '').strip() for pk in request.POST.getlist('idiomas')
        }
        return self.render_to_response(self.get_context_data(
            form=form, form_personales=form_personales, form_complementarios=form_complementarios,
            discapacidades=self._checkboxes(Discapacidad.objects.all(), set(request.POST.getlist('discapacidades'))),
            enfermedades=self._checkboxes(EnfermedadCronica.objects.all(), set(request.POST.getlist('enfermedades'))),
            idiomas=self._checkboxes(Idioma.objects.all(), set(request.POST.getlist('idiomas')), niveles=niveles_post),
        ))


@login_required
def servidor_baja(request, pk):
    qs = filtrar_por_dependencia(
        ServidorPublico.objects.all(), request.user, 'informacion_basica__dependencia_id'
    ).distinct()
    servidor = get_object_or_404(qs, pk=pk)
    if request.method == 'POST':
        form = BajaForm(request.POST)
        if not request.user.es_administrador:
            form.fields['dependencia'].queryset = form.fields['dependencia'].queryset.filter(
                pk=request.user.dependencia_id
            )
        if form.is_valid():
            if not request.user.es_administrador and form.cleaned_data['dependencia'].pk != request.user.dependencia_id:
                form.add_error('dependencia', 'No tiene permiso para asignar esta dependencia.')
            else:
                baja = form.save(commit=False)
                baja.servidor = servidor
                baja.registrado_por = request.user
                baja.save()
                LogEvento.registrar('baja', baja.pk, servidor, 'creado', 'manual', usuario=request.user)
                servidor.activo = False
                servidor.save()
                InformacionBasica.objects.filter(servidor=servidor, activo=True).update(activo=False)
                liberar_puestos_de(servidor)
                return redirect('servidor_list')
    else:
        form = BajaForm()
        if not request.user.es_administrador:
            form.fields['dependencia'].queryset = form.fields['dependencia'].queryset.filter(
                pk=request.user.dependencia_id
            )
            form.fields['dependencia'].initial = request.user.dependencia_id
    return render(request, 'servidores/baja_form.html', {
        'form': form, 'servidor': servidor, 'titulo': f'Registrar Baja: {servidor.nombre_completo}'
    })


class BajaUpdateView(LoginRequiredMixin, DependenciaScopedMixin, UpdateView):
    """Corrige los datos administrativos de una baja ya registrada (por
    layout o manual). No permite cambiar el servidor ni la plaza liberada
    —eso ya tuvo su efecto al darla de alta— solo dependencia/fecha/
    motivo/ejercicio/periodo, igual que el alta manual (BajaForm)."""
    model = BajaServidorPublico
    form_class = BajaForm
    template_name = 'servidores/baja_edit_form.html'
    success_url = reverse_lazy('reporte_bajas')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f'Editar Baja: {self.object.servidor.nombre_completo}'
        return ctx

    def form_valid(self, form):
        response = super().form_valid(form)
        LogEvento.registrar('baja', self.object.pk, self.object.servidor, 'editado', 'manual', usuario=self.request.user)
        return response


@login_required
def hoja_resp(request, pk):
    qs = filtrar_por_dependencia(
        ServidorPublico.objects.all(), request.user, 'informacion_basica__dependencia_id'
    ).distinct()
    servidor = get_object_or_404(qs, pk=pk)
    info = InformacionBasica.objects.filter(servidor=servidor, activo=True).order_by('-quincena').first()
    return render(request, 'servidores/hoja_resp.html', {
        'servidor': servidor, 'info': info, 'titulo': 'Hoja RESP'
    })


class InformacionBasicaListView(LoginRequiredMixin, PermisoRequeridoMixin, ListView):
    """Por default solo muestra lo VIGENTE (activo=True) — lo mismo que
    alimenta reportes y estadísticas. Con ?vista=todas se ve el histórico
    completo (todas las quincenas, no solo la vigente por servidor+plaza),
    útil para consultar/corregir una carga anterior; cada registro con
    carga_origen enlaza de vuelta a esa carga. Con ?carga=<pk> se filtra a
    solo lo generado por esa carga en particular (link desde su detalle)."""
    permiso_clave = 'info_basica'
    model = InformacionBasica
    template_name = 'servidores/info_basica_list.html'
    context_object_name = 'registros'
    paginate_by = 25

    def get_queryset(self):
        qs = filtrar_por_dependencia(
            InformacionBasica.objects.select_related('servidor', 'dependencia', 'estatus_plaza', 'carga_origen'),
            self.request.user,
        )
        carga_pk = self.request.GET.get('carga', '')
        if carga_pk:
            # Ver lo que generó ESA carga en particular importa sin importar
            # si sigue vigente o ya quedó reemplazada por una carga posterior.
            qs = qs.filter(carga_origen_id=carga_pk)
        elif self.request.GET.get('vista') != 'todas':
            qs = qs.filter(activo=True)
        q = self.request.GET.get('q', '').strip()
        if q:
            # Búsqueda por nombre completo (paterno + materno + nombre) sin
            # importar el orden en que se escriban los pedazos: cada palabra
            # de 'q' se busca por separado (AND) en el nombre concatenado,
            # así "lope elva", "lop cad elva" o solo "elva" encuentran a
            # "LOPEZ CADENA ELVA" igual. RFC/ID de plaza siguen comparándose
            # contra 'q' completo, como antes.
            qs = qs.annotate(
                nombre_busqueda=Concat(
                    'servidor__primer_apellido', Value(' '),
                    'servidor__segundo_apellido', Value(' '),
                    'servidor__nombre',
                )
            )
            filtro_nombre = Q()
            for token in q.split():
                filtro_nombre &= Q(nombre_busqueda__icontains=token)
            qs = qs.filter(filtro_nombre | Q(servidor__rfc__icontains=q) | Q(id_plaza__icontains=q))
        return qs.order_by('-quincena', 'servidor__rfc')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Información Básica'
        ctx['vista'] = self.request.GET.get('vista', 'vigentes')
        ctx['carga_filtro'] = self.request.GET.get('carga', '')
        return ctx


ICONOS_ESTATUS_PLAZA = {
    'ocupada':     ('👤', 'verde'),
    'vacante':     ('🪑', 'rojo'),
    'licencia':    ('🌴', 'dorado'),
    'comisionado': ('🔄', 'azul'),
    'reservada':   ('🔒', 'azul'),
    'suspendido':  ('⛔', 'rojo'),
}


class PuestoListView(LoginRequiredMixin, PermisoRequeridoMixin, ListView):
    permiso_clave = 'plazas'
    model = Puesto
    template_name = 'servidores/puesto_list.html'
    context_object_name = 'puestos'
    paginate_by = 25

    def get_queryset(self):
        qs = filtrar_por_dependencia(
            Puesto.objects.select_related(
                'proyecto', 'proyecto__dependencia', 'programa', 'programa__unidad',
                'unidad', 'categoria', 'servidor_actual', 'estatus_plaza'
            ),
            self.request.user, 'proyecto__dependencia',
        )
        q = self.request.GET.get('q', '')
        if q:
            qs = qs.filter(
                Q(id_plaza__icontains=q) | Q(proyecto__clave__icontains=q) |
                Q(proyecto__dependencia__clave__icontains=q) |
                Q(servidor_actual__rfc__icontains=q)
            )
        estatus = self.request.GET.get('estatus', '')
        if estatus:
            qs = qs.filter(estatus_plaza_id=estatus)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Plazas'
        ctx['q'] = self.request.GET.get('q', '')
        ctx['estatus'] = self.request.GET.get('estatus', '')
        ctx['estatus_choices'] = EstatusPlaza.objects.all()

        puestos_visibles = filtrar_por_dependencia(Puesto.objects.all(), self.request.user, 'proyecto__dependencia')
        resumen = list(
            puestos_visibles.values('estatus_plaza__descripcion').annotate(total=Count('id')).order_by('-total')
        )
        for item in resumen:
            item['descripcion'] = item['estatus_plaza__descripcion'] or 'Sin estatus'
            item['icono'], item['color'] = ICONOS_ESTATUS_PLAZA.get(item['descripcion'].strip().lower(), ('📌', 'azul'))
        ctx['resumen_estatus'] = resumen
        ctx['total_plazas'] = puestos_visibles.count()
        return ctx


PLAZAS_EXPORT_HEADERS = [
    'FUENTE FINANCIAMIENTO', 'DEPENDENCIA', 'UNIDAD', 'PROGRAMA', 'PROYECTO', 'CATEGORIA',
    'TIPO CONTRATACION', 'TIPO PERSONAL', 'TIPO FUNCION', 'NIVEL ESTRUCTURA', 'ID PUESTO',
    'ID PUESTO JEFE', 'HSM', 'PERCEPCIONES', 'BONOS', 'NETO', 'DIAS PAGADOS', 'ESTATUS PLAZA',
    'EXPEDIENTE', 'RFC', 'CURP', 'DETERMINANTE', 'NOMBRE', 'PRIMER APELLIDO', 'SEGUNDO APELLIDO',
    'FECHA NACIMIENTO', 'GENERO', 'ESTADO CIVIL', 'ENTIDAD', 'PAIS', 'CORREO ELECTRONICO', 'ISS',
    'NSS', 'CENTRO TRABAJO', 'SINDICALIZADO', 'SINDICATO', 'ORDP', 'TIPO DECLARACION', 'OPAAER',
    'RENCTA', 'PRDMIS', 'OTRA PLAZA', 'FECHA INGRESO GOBIERNO', 'FECHA INGRESO DEPENDENCIA',
    'FECHA INGRESO PUESTO', 'AREA', 'PARCONPUB', 'CONTRATACION PUBLICA', 'PARCON', 'CONCESIONES',
    'PARENA', 'ENAJENACION', 'INMUEBLE',
]


@login_required
def exportar_plazas_excel(request):
    """Exporta las plazas en el mismo formato de 53 columnas que el Layout de
    Información Básica (mismo orden que cargas/procesador.py: C_FTE_FINAN..
    C_INMUEBLE) — administradores ven/exportan todas, el resto solo las de
    su propia dependencia. La mayoría de las columnas de 'Persona-Puesto',
    'Fechas' y 'Responsabilidades' no viven en Puesto ni en ServidorPublico
    (solo en InformacionBasica, sin FK directa a Puesto), así que se buscan
    por el último registro de InformacionBasica de ese servidor+plaza; en
    una plaza vacante, o sin InformacionBasica todavía, esas columnas quedan
    vacías — igual que se verían en una fila de "reporte de vacante" del
    layout real."""
    puestos = filtrar_por_dependencia(
        Puesto.objects.select_related(
            'proyecto', 'proyecto__dependencia', 'programa', 'unidad', 'categoria',
            'nombramiento', 'nivel_estructura', 'estatus_plaza', 'cct',
            'servidor_actual', 'servidor_actual__estado_civil', 'servidor_actual__entidad_nacimiento',
            'servidor_actual__pais_nacimiento', 'servidor_actual__sindicato',
        ),
        request.user, 'proyecto__dependencia',
    ).order_by('proyecto__dependencia__clave', 'id_plaza')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Plazas'
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(PLAZAS_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    ib_cache = {}

    def info_basica_de(servidor, id_plaza):
        if not servidor:
            return None
        clave = (servidor.pk, id_plaza)
        if clave not in ib_cache:
            ib_cache[clave] = InformacionBasica.objects.filter(
                servidor=servidor, id_plaza=id_plaza,
            ).select_related(
                'fuente_financiamiento', 'tipo_personal', 'tipo_funcion', 'tipo_declaracion',
                'area', 'inmueble',
            ).order_by('-activo', '-quincena').first()
        return ib_cache[clave]

    row = 2
    for p in puestos:
        s = p.servidor_actual
        ib = info_basica_de(s, p.id_plaza)
        valores = [
            ib.fuente_financiamiento.clave if ib and ib.fuente_financiamiento else '',
            p.proyecto.dependencia.clave if p.proyecto and p.proyecto.dependencia else '',
            p.unidad.clave if p.unidad else '',
            p.programa.clave if p.programa else '',
            p.proyecto.clave if p.proyecto else '',
            p.categoria.clave if p.categoria else '',
            p.nombramiento.clave if p.nombramiento else '',
            ib.tipo_personal.clave if ib and ib.tipo_personal else '',
            ib.tipo_funcion.clave if ib and ib.tipo_funcion else '',
            p.nivel_estructura.nivel if p.nivel_estructura else '',
            p.id_plaza,
            p.id_plaza_jefe,
            float(p.hsm) if p.hsm is not None else '',
            float(p.total_percepciones),
            float(p.total_bonos),
            float(p.total_neto),
            p.dias_pagados,
            p.estatus_plaza.clave if p.estatus_plaza else '',
            s.expediente if s else '',
            s.rfc if s else '',
            s.curp if s else '',
            s.determinante if s else '',
            s.nombre if s else '',
            s.primer_apellido if s else '',
            s.segundo_apellido if s else '',
            s.fecha_nacimiento if s else '',
            s.sexo if s else '',
            s.estado_civil.clave if s and s.estado_civil else '',
            s.entidad_nacimiento.clave if s and s.entidad_nacimiento else '',
            s.pais_nacimiento.clave if s and s.pais_nacimiento else '',
            s.correo_institucional if s else '',
            s.iss if s else '',
            s.nss if s else '',
            p.cct.clave if p.cct else '',
            s.sindicalizado if s else '',
            s.sindicato.clave if s and s.sindicato else '',
            ib.oblig_declaracion if ib else '',
            ib.tipo_declaracion.clave if ib and ib.tipo_declaracion else '',
            ib.oblig_entrega_recepcion if ib else '',
            ib.oblig_rendir_cuentas if ib else '',
            ib.puesto_sensible if ib else '',
            s.tiene_otra_plaza if s else '',
            ib.fecha_ingreso_gobierno if ib else '',
            ib.fecha_ingreso_dependencia if ib else '',
            ib.fecha_ingreso_puesto if ib else '',
            ib.area.clave if ib and ib.area else '',
            ib.participa_contrataciones if ib else '',
            ib.nivel_contrataciones if ib else '',
            ib.participa_concesiones if ib else '',
            ib.nivel_concesiones if ib else '',
            ib.participa_enajenacion if ib else '',
            ib.nivel_enajenacion if ib else '',
            ib.inmueble.clave if ib and ib.inmueble else '',
        ]
        for col, valor in enumerate(valores, 1):
            ws.cell(row=row, column=col, value=valor)
        row += 1

    for i in range(1, len(PLAZAS_EXPORT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="RESP_Plazas.xlsx"'
    wb.save(response)
    return response


class PuestoDetailView(LoginRequiredMixin, DependenciaScopedMixin, DetailView):
    """Detalle de una plaza: quién la ocupa hoy, y su historial — igual que
    el detalle de un servidor, pero visto desde el lado de la plaza: qué
    servidores la han ocupado quincena a quincena (InformacionBasica por
    id_plaza) y qué bajas la han liberado (BajaServidorPublico por id_plaza,
    ver cargas/procesador.py: procesar_layout_bajas ya guarda ese vínculo)."""
    model = Puesto
    template_name = 'servidores/puesto_detail.html'
    context_object_name = 'puesto'
    dependencia_lookup = 'proyecto__dependencia'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['historial'] = InformacionBasica.objects.filter(
            id_plaza=self.object.id_plaza
        ).select_related('servidor', 'dependencia', 'nombramiento', 'estatus_plaza').order_by('-quincena')[:15]
        ctx['bajas'] = BajaServidorPublico.objects.filter(
            id_plaza=self.object.id_plaza
        ).select_related('servidor', 'motivo_baja', 'dependencia').order_by('-fecha_baja')
        ctx['titulo'] = f'Plaza: {self.object.id_plaza}'
        return ctx


class PuestoUpdateView(LoginRequiredMixin, DependenciaScopedMixin, UpdateView):
    model = Puesto
    form_class = PuestoForm
    template_name = 'servidores/puesto_form.html'
    success_url = reverse_lazy('puesto_list')
    dependencia_lookup = 'proyecto__dependencia'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f'Modificar Plaza: {self.object.id_plaza}'
        return ctx

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        user = self.request.user
        if not user.es_administrador:
            form.fields['dependencia'].queryset = form.fields['dependencia'].queryset.filter(pk=user.dependencia_id)
            form.fields['dependencia'].initial = user.dependencia_id
            for campo in ('proyecto', 'programa', 'unidad'):
                form.fields[campo].queryset = form.fields[campo].queryset.filter(dependencia_id=user.dependencia_id)
        return form

    def form_valid(self, form):
        user = self.request.user
        if not user.es_administrador:
            proyecto = form.cleaned_data.get('proyecto')
            if proyecto and proyecto.dependencia_id != user.dependencia_id:
                form.add_error('proyecto', 'No tiene permiso para asignar un proyecto de otra dependencia.')
                return self.form_invalid(form)
        return super().form_valid(form)


class InformacionBasicaCreateView(LoginRequiredMixin, CreateView):
    """No usa DependenciaFormRestrictMixin: su form_valid() ya está sobrescrito
    para sincronizar la Plaza tras guardar, y esa sincronización no debe
    ejecutarse si la validación de dependencia rechaza el POST. Se valida
    explícito al inicio de este mismo form_valid en su lugar."""
    model = InformacionBasica
    form_class = InformacionBasicaForm
    template_name = 'servidores/info_basica_form.html'
    success_url = reverse_lazy('info_basica_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        user = self.request.user
        if 'dependencia' in form.fields and not user.es_administrador:
            form.fields['dependencia'].queryset = form.fields['dependencia'].queryset.filter(pk=user.dependencia_id)
            form.fields['dependencia'].initial = user.dependencia_id
        return form

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Nueva Información Básica'
        return ctx

    def form_valid(self, form):
        user = self.request.user
        if not user.es_administrador:
            dep = form.cleaned_data.get('dependencia')
            if dep and dep.pk != user.dependencia_id:
                form.add_error('dependencia', 'No tiene permiso para asignar esta dependencia.')
                return self.form_invalid(form)
        response = super().form_valid(form)
        info = self.object
        sincronizar_puesto(
            info.proyecto, info.programa, info.id_plaza, info.categoria, info.servidor,
            unidad=info.unidad,
            nombramiento=info.nombramiento,
            nivel_estructura=info.nivel_estructura,
            estatus_plaza=info.estatus_plaza,
            cct=info.cct,
            hsm=info.hsm,
            total_percepciones=info.total_percepciones,
            total_bonos=info.total_bonos,
            total_neto=info.total_neto,
            dias_pagados=info.dias_pagados,
            id_plaza_jefe=info.id_plaza_jefe,
        )
        LogEvento.registrar('informacion_basica', info.pk, info.servidor, 'creado', 'manual', usuario=user)
        return response


class InformacionBasicaUpdateView(LoginRequiredMixin, DependenciaScopedMixin, UpdateView):
    model = InformacionBasica
    form_class = InformacionBasicaForm
    template_name = 'servidores/info_basica_form.html'
    success_url = reverse_lazy('info_basica_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        user = self.request.user
        if 'dependencia' in form.fields and not user.es_administrador:
            form.fields['dependencia'].queryset = form.fields['dependencia'].queryset.filter(pk=user.dependencia_id)
            form.fields['dependencia'].initial = user.dependencia_id
        return form

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = 'Modificar Información Básica'
        return ctx

    def form_valid(self, form):
        user = self.request.user
        if not user.es_administrador:
            dep = form.cleaned_data.get('dependencia')
            if dep and dep.pk != user.dependencia_id:
                form.add_error('dependencia', 'No tiene permiso para asignar esta dependencia.')
                return self.form_invalid(form)
        response = super().form_valid(form)
        info = self.object
        sincronizar_puesto(
            info.proyecto, info.programa, info.id_plaza, info.categoria, info.servidor,
            unidad=info.unidad,
            nombramiento=info.nombramiento,
            nivel_estructura=info.nivel_estructura,
            estatus_plaza=info.estatus_plaza,
            cct=info.cct,
            hsm=info.hsm,
            total_percepciones=info.total_percepciones,
            total_bonos=info.total_bonos,
            total_neto=info.total_neto,
            dias_pagados=info.dias_pagados,
            id_plaza_jefe=info.id_plaza_jefe,
        )
        LogEvento.registrar('informacion_basica', info.pk, info.servidor, 'editado', 'manual', usuario=user)
        return response
