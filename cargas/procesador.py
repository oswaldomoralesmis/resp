# -*- coding: utf-8 -*-
"""
Procesador de layouts para el RESP.

Columnas del Layout Básico (orden exacto del archivo oficial):
 0  FUENTE FINANCIAMIENTO
 1  DEPENDENCIA
 2  UNIDAD
 3  PROGRAMA
 4  PROYECTO
 5  CATEGORIA
 6  TIPO CONTRATACION
 7  TIPO PERSONAL
 8  TIPO FUNCION
 9  NIVEL ESTRUCTURA
10  ID PUESTO
11  ID PUESTO JEFE
12  HSM
13  PERCEPCIONES
14  BONOS
15  NETO
16  DIAS PAGADOS
17  ESTATUS PLAZA
18  EXPEDIENTE
19  RFC
20  CURP
21  DETERMINANTE
22  NOMBRE
23  PRIMER APELLIDO
24  SEGUNDO APELLIDO
25  FECHA NACIMIENTO
26  GENERO
27  ESTADO CIVIL
28  ENTIDAD
29  PAIS
30  CORREO ELECTRONICO
31  ISS
32  NSS
33  CENTRO TRABAJO
34  SINDICALIZADO
35  SINDICATO
36  ORDP               (Obligado Declaracion Patrimonial S/N/NULL)
37  TIPO DECLARACION
38  OPAAER             (Obligado Presentar Acta Entrega-Recepcion S/N/NULL)
39  RENCTA             (Rinde Cuentas S/N/NULL)
40  PRDMIS             (Puesto Responsabilidad Decision / Manejo Info Sensible S/N/NULL)
41  OTRA PLAZA
42  FECHA INGRESO GOBIERNO
43  FECHA INGRESO DEPENDENCIA
44  FECHA INGRESO PUESTO
45  AREA
46  PARCONPUB          (Participa Contrataciones Publicas S/N)
47  CONTRATACION PUBLICA (nivel A/B/C/NULL)
48  PARCON             (Participa Concesiones S/N)
49  CONCESIONES        (nivel A/B/C/NULL)
50  PARENA             (Participa Enajenacion S/N)
51  ENAJENACION        (nivel A/B/C/NULL)
52  INMUEBLE           (clave inmueble)
"""
import logging
from datetime import datetime, date

from django.db import transaction
from openpyxl import load_workbook

from catalogos.models import (
    FuenteFinanciamiento, Dependencia, UnidadAdministrativa,
    Programa, Proyecto, Categoria, TipoContratacion, TipoPersonal,
    TipoFuncion, NivelEstructura, EstatusPlaza, CentroTrabajo,
    TipoDeclaracion, Area, EntidadFederativa, Pais, Inmueble,
    EstadoCivil, Sindicato, MotivoBaja,
)
from .validadores import rfc_formato_valido, curp_formato_valido, curp_digito_verificador_valido
from servidores.models import (
    ServidorPublico, InformacionBasica, Puesto, BajaServidorPublico,
    sincronizar_puesto, reportar_puesto_vacante,
)

logger = logging.getLogger(__name__)

# ── Índices de columna (base 0) ───────────────────────────────────────────────
C_FTE_FINAN      = 0
C_DEPENDENCIA    = 1
C_UNIDAD         = 2
C_PROGRAMA       = 3
C_PROYECTO       = 4
C_CATEGORIA      = 5
C_TIPO_CONTRA    = 6
C_TIPO_PERSONAL  = 7
C_TIPO_FUNCION   = 8
C_NIVEL_EST      = 9
C_ID_PUESTO      = 10
C_ID_PUESTO_JEFE = 11
C_HSM            = 12
C_PERCEPCIONES   = 13
C_BONOS          = 14
C_NETO           = 15
C_DIAS_PAGADOS   = 16
C_ESTATUS_PLAZA  = 17
C_EXPEDIENTE     = 18
C_RFC            = 19
C_CURP           = 20
C_DETERMINANTE   = 21
C_NOMBRE         = 22
C_PRIMER_AP      = 23
C_SEGUNDO_AP     = 24
C_FECHA_NAC      = 25
C_GENERO         = 26
C_ESTADO_CIVIL   = 27
C_ENTIDAD        = 28
C_PAIS           = 29
C_CORREO         = 30
C_ISS            = 31
C_NSS            = 32
C_CCT            = 33
C_SINDICALIZADO  = 34
C_SINDICATO      = 35
C_ORDP           = 36   # Obligado Declaracion Patrimonial
C_TIPO_DECLA     = 37
C_OPAAER         = 38   # Obligado Entrega-Recepcion
C_RENCTA         = 39   # Rinde Cuentas
C_PRDMIS         = 40   # Puesto Resp. Decision / Info Sensible
C_OTRA_PLAZA     = 41
C_F_INGOB        = 42   # Fecha Ingreso Gobierno
C_F_INDEP        = 43   # Fecha Ingreso Dependencia
C_F_INPUES       = 44   # Fecha Ingreso Puesto
C_AREA           = 45
C_PARCONPUB      = 46   # Participa Contrataciones
C_CONTRAT_PUB    = 47   # Nivel Contrataciones
C_PARCON         = 48   # Participa Concesiones
C_CONCESIONES    = 49   # Nivel Concesiones
C_PARENA         = 50   # Participa Enajenacion
C_ENAJENACION    = 51   # Nivel Enajenacion
C_INMUEBLE       = 52

# Errores "forzables": tienen un valor capturado (no vacío) que solo falla una
# validación de formato — el validador puede, fila por fila y con motivo,
# aceptarlos tal cual vinieron. El resto de los errores (campo vacío, plaza/
# servidor/motivo inexistente) no son forzables: no hay dato válido que
# escribir, o forzarlos violaría otra regla de negocio ya definida.
MENSAJES_FORZABLES = {'RFC con formato inválido', 'CURP con formato inválido'}


def es_fila_forzable(errores_fila):
    """True si TODOS los problemas de la fila son forzables (si hay aunque
    sea uno que no lo es, la fila sigue bloqueada sin importar el override)."""
    return bool(errores_fila) and all(e in MENSAJES_FORZABLES for e in errores_fila)


# ── Comparación contra el período anterior (aviso, nunca bloquea) ─────────────
# Todas las columnas de InformacionBasica que vienen del layout básico, para
# poder avisarle al validador exactamente qué cambió respecto a la última
# quincena cargada de ese mismo servidor+plaza, antes de aceptar la carga.
CAMPOS_COMPARABLES_BASICA = [
    ('fuente_financiamiento',    'Fuente de Financiamiento'),
    ('unidad',                   'Unidad Administrativa'),
    ('programa',                 'Programa'),
    ('proyecto',                 'Proyecto'),
    ('categoria',                'Categoría'),
    ('nombramiento',             'Tipo de Contratación'),
    ('tipo_personal',            'Tipo de Personal'),
    ('tipo_funcion',             'Tipo de Función'),
    ('nivel_estructura',         'Nivel de Estructura'),
    ('id_plaza_jefe',            'ID Plaza Jefe'),
    ('hsm',                      'HSM'),
    ('total_percepciones',       'Percepciones'),
    ('total_bonos',              'Bonos'),
    ('total_neto',               'Neto'),
    ('dias_pagados',             'Días Pagados'),
    ('estatus_plaza',            'Estatus de Plaza'),
    ('cct',                      'Centro de Trabajo'),
    ('oblig_declaracion',        'Obligado Declaración Patrimonial'),
    ('tipo_declaracion',         'Tipo de Declaración'),
    ('oblig_entrega_recepcion',  'Oblig. Entrega-Recepción'),
    ('oblig_rendir_cuentas',     'Obligado Rendir Cuentas'),
    ('puesto_sensible',          'Puesto Sensible'),
    ('fecha_ingreso_gobierno',   'Fecha Ingreso Gobierno'),
    ('fecha_ingreso_dependencia','Fecha Ingreso Dependencia'),
    ('fecha_ingreso_puesto',     'Fecha Ingreso Puesto'),
    ('area',                     'Área'),
    ('participa_contrataciones', 'Participa Contrataciones'),
    ('nivel_contrataciones',     'Nivel Contrataciones'),
    ('participa_concesiones',    'Participa Concesiones'),
    ('nivel_concesiones',        'Nivel Concesiones'),
    ('participa_enajenacion',    'Participa Enajenación'),
    ('nivel_enajenacion',        'Nivel Enajenación'),
    ('inmueble',                 'Inmueble'),
]
_CAMPOS_NUMERICOS_BASICA = {'hsm', 'total_percepciones', 'total_bonos', 'total_neto'}


def comparar_con_periodo_anterior(servidor, id_plaza, quincena_actual, valores_nuevos):
    """Compara 'valores_nuevos' (lo que se va a escribir en InformacionBasica
    para esta fila) contra el último registro que ya existía en BD para ese
    mismo servidor+plaza, en una quincena distinta a la actual. Devuelve una
    lista de descripciones de cambio por columna (para avisos — nunca bloquea
    la fila), o [] si no hay período anterior con qué comparar (alta nueva)."""
    previo = InformacionBasica.objects.filter(
        servidor=servidor, id_plaza=id_plaza,
    ).exclude(quincena=quincena_actual).order_by('-quincena').first()
    if not previo:
        return []
    cambios = []
    for campo, etiqueta in CAMPOS_COMPARABLES_BASICA:
        valor_previo = getattr(previo, campo)
        valor_nuevo = valores_nuevos.get(campo)
        if campo in _CAMPOS_NUMERICOS_BASICA:
            try:
                iguales = float(valor_previo or 0) == float(valor_nuevo or 0)
            except (TypeError, ValueError):
                iguales = valor_previo == valor_nuevo
        else:
            iguales = valor_previo == valor_nuevo
        if not iguales:
            v_previo = valor_previo if valor_previo not in (None, '') else '—'
            v_nuevo = valor_nuevo if valor_nuevo not in (None, '') else '—'
            cambios.append(f'{etiqueta} cambió de "{v_previo}" a "{v_nuevo}" (vs. quincena {previo.quincena})')
    return cambios


# ── Helpers ───────────────────────────────────────────────────────────────────
def sv(v, default=''):
    """Safe string, fuerza mayúsculas, limpia NULL."""
    if v is None:
        return default
    s = str(v).strip().upper()
    return s if s not in ('NULL', 'NONE', 'N/A', '') else default


def sd(v, default=''):
    """Safe string, preserva case."""
    if v is None:
        return default
    s = str(v).strip()
    return s if s.upper() not in ('NULL', 'NONE', 'N/A', '') else default


def sf(v, default=0.0):
    try:
        return float(v) if v is not None else default
    except (ValueError, TypeError):
        return default


def si(v, default=0):
    try:
        return int(float(str(v))) if v is not None else default
    except (ValueError, TypeError):
        return default


def parse_fecha(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def normalizar_sexo(v):
    s = sv(v)
    if s in ('M', 'MASCULINO', 'H', 'HOMBRE', 'MALE', '1'):
        return 'MASCULINO'
    if s in ('F', 'FEMENINO', 'MUJER', 'FEMALE', '2'):
        return 'FEMENINO'
    return 'OTRO'


def normalizar_sino(v):
    s = sv(v)
    if s in ('S', 'SI', 'YES', '1', 'TRUE'):
        return 'S'
    if s in ('N', 'NO', '0', 'FALSE'):
        return 'N'
    return 'NULL'


def normalizar_iss(v):
    s = sv(v)
    if 'ISSET' in s:
        return 'ISSET'
    if 'IMSS' in s:
        return 'IMSS'
    if 'ISSSTE' in s:
        return 'ISSSTE'
    return 'ISSET'


# ── Caché de catálogos ────────────────────────────────────────────────────────
class Cache:
    def __init__(self):
        self._d = {}

    def _get(self, key, fn):
        if key not in self._d:
            self._d[key] = fn()
        return self._d[key]

    def fuente(self, v):
        k = sv(v)
        return self._get(f'fte_{k}', lambda: FuenteFinanciamiento.objects.filter(clave=k).first())

    def dependencia(self, v):
        k = sv(v)
        return self._get(f'dep_{k}', lambda: Dependencia.objects.filter(clave=k).first())

    def unidad(self, v):
        k = sv(v)
        return self._get(f'uni_{k}', lambda: UnidadAdministrativa.objects.filter(clave=k).first())

    def programa(self, v):
        k = sv(v)
        return self._get(f'prg_{k}', lambda: Programa.objects.filter(clave=k).first())

    def proyecto(self, v, dependencia=None):
        k = sd(v).strip()
        key = f'pry_{k}_{dependencia.pk if dependencia else ""}'

        def _buscar():
            qs = Proyecto.objects.filter(clave__iexact=k)
            # Un proyecto es único por dependencia-clave; puede abarcar
            # varios programas/unidades (Proyecto.programas es M2M).
            if dependencia:
                pry = qs.filter(dependencia=dependencia).first()
                if pry:
                    return pry
            pry = qs.first()
            if pry:
                return pry
            # Si la clave es 00000000 o no se encontró, usar el proyecto default
            if k in ('00000000', '', '0'):
                # Proyecto default vinculado a la dependencia
                if dependencia:
                    pry_def = Proyecto.objects.filter(
                        clave='00000000',
                        dependencia=dependencia
                    ).first()
                    if pry_def:
                        return pry_def
                # Cualquier proyecto default
                return Proyecto.objects.filter(clave='00000000').first()
            return None

        return self._get(key, _buscar)

    def categoria(self, v):
        k = sv(v)
        return self._get(f'cat_{k}', lambda: Categoria.objects.filter(clave=k).first())

    def tipo_contratacion(self, v):
        k = sv(v)
        return self._get(f'tc_{k}', lambda: (
            TipoContratacion.objects.filter(clave=k).first() or
            TipoContratacion.objects.filter(descripcion__icontains=k).first()
        ))

    def tipo_personal(self, v):
        k = sv(v)
        return self._get(f'tp_{k}', lambda: TipoPersonal.objects.filter(clave=k).first())

    def tipo_funcion(self, v):
        k = sv(v)
        return self._get(f'tf_{k}', lambda: TipoFuncion.objects.filter(clave=k).first())

    def nivel_estructura(self, v):
        k = si(v)
        return self._get(f'ne_{k}', lambda: NivelEstructura.objects.filter(nivel=k).first())

    def estatus_plaza(self, v):
        k = sv(v)
        return self._get(f'ep_{k}', lambda: (
            EstatusPlaza.objects.filter(clave=k).first() or
            EstatusPlaza.objects.filter(descripcion__icontains=k).first()
        ))

    def centro_trabajo(self, v):
        k = sd(v)
        return self._get(f'cct_{k}', lambda: CentroTrabajo.objects.filter(clave__iexact=k).first())

    def tipo_declaracion(self, v):
        k = sv(v)
        return self._get(f'td_{k}', lambda: (
            TipoDeclaracion.objects.filter(clave=k).first() or
            TipoDeclaracion.objects.filter(descripcion__icontains=k).first()
        ))

    def area(self, v):
        k = sv(v)
        return self._get(f'ar_{k}', lambda: Area.objects.filter(clave=k).first())

    def entidad(self, v):
        k = sv(v)
        return self._get(f'ent_{k}', lambda: (
            EntidadFederativa.objects.filter(clave=si(v)).first() if str(v).strip().isdigit()
            else EntidadFederativa.objects.filter(nombre__icontains=k).first()
        ))

    def pais(self, v):
        k = sv(v)
        return self._get(f'pai_{k}', lambda: (
            Pais.objects.filter(clave=si(v)).first() if str(v).strip().isdigit()
            else Pais.objects.filter(nombre__icontains=k).first()
            or Pais.objects.filter(clave=700).first()
        ))

    def estado_civil(self, v):
        k = sv(v)
        return self._get(f'ec_{k}', lambda: EstadoCivil.objects.filter(clave=k).first())

    def sindicato(self, v):
        k = sv(v)
        return self._get(f'sin_{k}', lambda: Sindicato.objects.filter(clave=k).first())

    def inmueble(self, v):
        k = sd(v)
        if not k:
            return None
        return self._get(f'inm_{k}', lambda: Inmueble.objects.filter(clave__iexact=k).first())

    def motivo_baja(self, v):
        k = sv(v)
        return self._get(f'mb_{k}', lambda: MotivoBaja.objects.filter(clave=k).first())


# ── Procesador principal ──────────────────────────────────────────────────────
def procesar_layout_basica(carga, dry_run=True, overrides=None):
    """Si dry_run=True (vista previa al momento de la carga), valida cada fila
    y arma el mismo log de siempre, pero SIN escribir en ServidorPublico/Puesto/
    InformacionBasica. Si dry_run=False (al aceptar la carga), aplica los
    cambios de verdad. En ambos casos el resultado tiene el mismo formato.

    'overrides' (dict {num_fila: {'decision': 'aceptada'|'rechazada', 'motivo': str}})
    son las decisiones que el validador tomó en la pantalla de revisión sobre
    filas con error FORZABLE (ver MENSAJES_FORZABLES) — únicamente errores de
    formato de RFC/CURP, donde sí hay un valor capturado y forzarlo no implica
    inventar datos ni violar la regla de "las plazas no se crean desde el
    layout". Si una fila forzable fue 'aceptada', esa validación puntual se
    salta y se registra como aviso (con el motivo del validador) en vez de
    bloquear la fila."""
    overrides = overrides or {}
    ruta     = carga.archivo.path
    # InformacionBasica.quincena es un identificador histórico (puede acumular
    # varios ejercicios por servidor), por eso aquí sí se guarda con el año.
    quincena = f'{carga.periodo.ejercicio}-{carga.periodo.quincena}'
    log      = []
    filas_detalle = []  # listado completo de registros para la revisión del validador
    ok       = 0
    errores  = 0
    total    = 0

    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
    except Exception as e:
        return {'ok': 0, 'errores': 1, 'total': 0, 'log': f'No se pudo abrir el archivo: {e}', 'filas': []}

    # Buscar la hoja correcta: primero "LAYOUT", luego la activa
    hoja = wb['LAYOUT'] if 'LAYOUT' in wb.sheetnames else wb.active
    filas = list(hoja.iter_rows(values_only=True))

    # Detectar la fila de encabezados buscando "RFC" o "CURP"
    inicio_datos = 1   # default: datos desde fila 2 (índice 1)
    for idx, fila in enumerate(filas):
        fila_str = [str(c).upper().strip() if c else '' for c in fila]
        if 'RFC' in fila_str or 'CURP' in fila_str:
            inicio_datos = idx + 1
            break

    cache = Cache()
    vistos_rfc = set()  # RFC ya "creados" en este mismo dry run (simula get_or_create sin escribir)

    for num_fila, fila in enumerate(filas[inicio_datos:], start=inicio_datos + 1):
        if not any(c is not None for c in fila):
            continue

        total += 1
        errores_fila = []

        def col(i, default=None):
            try:
                v = fila[i]
                if v is None or str(v).strip().upper() in ('NULL', 'NONE', 'N/A', ''):
                    return default
                return v
            except IndexError:
                return default

        # ── Campos obligatorios ──────────────────────────────────────────────
        rfc      = sv(col(C_RFC))
        curp     = sv(col(C_CURP))
        nombre   = sd(col(C_NOMBRE, '')).upper()
        ap_pat   = sd(col(C_PRIMER_AP, '')).upper()
        ap_mat   = sd(col(C_SEGUNDO_AP, '')).upper() or None
        id_plaza = sd(col(C_ID_PUESTO, ''))

        # ── Fila vacía: reporta una plaza sin trabajador → liberar el puesto ──
        # Las plazas NO se crean desde el layout (eso lo hace la carga inicial
        # de estructura); el id_plaza referido siempre debe existir ya.
        if not rfc and not curp and not nombre and not ap_pat and not ap_mat:
            if not id_plaza:
                errores += 1
                log.append(f'Fila {num_fila}: OMITIDA — fila vacía sin ID de plaza')
                filas_detalle.append({'fila': num_fila, 'rfc': '', 'curp': '', 'nombre': '',
                                       'id_plaza': '', 'estado': 'OMITIDA', 'mensaje': 'Fila vacía sin ID de plaza'})
                continue
            if not Puesto.objects.filter(id_plaza=id_plaza).exists():
                errores += 1
                log.append(f'Fila {num_fila}: ERROR — la plaza {id_plaza} no existe; debe darse de alta primero en la carga inicial de estructura')
                filas_detalle.append({'fila': num_fila, 'rfc': '', 'curp': '', 'nombre': '',
                                       'id_plaza': id_plaza, 'estado': 'ERROR',
                                       'mensaje': 'La plaza no existe; debe darse de alta primero en la carga inicial de estructura'})
                continue
            dependencia = cache.dependencia(col(C_DEPENDENCIA))
            if dependencia and dependencia.pk != carga.dependencia_id:
                errores += 1
                msg = (f'La dependencia de la fila ({dependencia.clave}) no coincide con '
                       f'la dependencia de la carga ({carga.dependencia.clave})')
                log.append(f'Fila {num_fila}: ERROR — {msg}')
                filas_detalle.append({'fila': num_fila, 'rfc': '', 'curp': '', 'nombre': '',
                                       'id_plaza': id_plaza, 'estado': 'ERROR', 'mensaje': msg})
                continue
            programa    = cache.programa(col(C_PROGRAMA))
            proyecto    = cache.proyecto(col(C_PROYECTO), dependencia=dependencia)
            categoria   = cache.categoria(col(C_CATEGORIA))
            if dry_run:
                ok += 1
                log.append(f'Fila {num_fila}: Plaza {id_plaza} sería reportada VACANTE.')
                filas_detalle.append({'fila': num_fila, 'rfc': '', 'curp': '', 'nombre': '',
                                       'id_plaza': id_plaza, 'estado': 'OK', 'mensaje': 'Plaza sería reportada VACANTE'})
                continue
            # La plaza ya existe (se validó arriba), así que reportar_puesto_vacante
            # solo la actualiza — nunca la crea — en este punto.
            puesto = reportar_puesto_vacante(id_plaza, proyecto=proyecto, programa=programa, categoria=categoria)
            ok += 1
            log.append(f'Fila {num_fila}: Plaza {id_plaza} reportada VACANTE.')
            filas_detalle.append({'fila': num_fila, 'rfc': '', 'curp': '', 'nombre': '',
                                   'id_plaza': id_plaza, 'estado': 'OK', 'mensaje': 'Plaza reportada VACANTE'})
            continue

        errores_fila_forzables = []
        if not rfc:
            errores_fila.append('RFC vacío')
        elif not rfc_formato_valido(rfc):
            errores_fila_forzables.append('RFC con formato inválido')
        if not curp:
            errores_fila.append('CURP vacío')
        elif not curp_formato_valido(curp):
            errores_fila_forzables.append('CURP con formato inválido')
        if not nombre:
            errores_fila.append('Nombre vacío')
        if not ap_pat and not ap_mat:
            # Basta con que venga alguno de los dos apellidos: hay servidores
            # que legítimamente solo tienen un apellido registrado.
            errores_fila.append('Apellido vacío (debe traer al menos uno: paterno o materno)')
        if not id_plaza:
            errores_fila.append('ID de plaza vacío')
        elif not Puesto.objects.filter(id_plaza=id_plaza).exists():
            errores_fila.append(f'La plaza {id_plaza} no existe (debe darse de alta primero en la carga inicial de estructura)')

        # Los errores forzables solo se pueden aceptar si son el ÚNICO problema
        # de la fila (si además falta un campo obligatorio, sigue bloqueada).
        avisos_forzados = []
        override = overrides.get(num_fila)
        if errores_fila_forzables:
            if not errores_fila and override and override.get('decision') == 'aceptada':
                motivo = override.get('motivo', '')
                avisos_forzados = [f'{msg} — ACEPTADO por el validador: {motivo}' for msg in errores_fila_forzables]
            else:
                errores_fila.extend(errores_fila_forzables)

        if errores_fila:
            errores += 1
            log.append(f'Fila {num_fila}: OMITIDA — {", ".join(errores_fila)}')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp,
                                   'nombre': f'{nombre} {ap_pat} {ap_mat or ""}'.strip(), 'id_plaza': id_plaza,
                                   'estado': 'OMITIDA', 'mensaje': ', '.join(errores_fila),
                                   'forzable': es_fila_forzable(errores_fila)})
            continue

        expediente = sd(col(C_EXPEDIENTE, '')) or f'EXP-{rfc}'

        # ── Servidor Público: buscar o crear ─────────────────────────────────
        # Si el servidor ya existía pero estaba inactivo (dado de baja), el
        # solo hecho de aparecer en un layout de Información Básica implica
        # que volvió a ocupar una plaza — se reactiva y se avisa.
        reactivado = False
        if dry_run:
            servidor = ServidorPublico.objects.filter(rfc=rfc).first()
            creado = servidor is None and rfc not in vistos_rfc
            if servidor is None:
                vistos_rfc.add(rfc)
                otro_por_curp = ServidorPublico.objects.filter(curp=curp).exclude(rfc=rfc).first()
                if otro_por_curp:
                    log.append(f'Fila {num_fila} RFC={rfc}: AVISO — CURP {curp} ya pertenece a otro RFC registrado')
            elif not servidor.activo:
                reactivado = True
        else:
            # Cada fila se guarda en su propio savepoint: si esta fila falla,
            # solo se revierte ella (no invalida el resto de la transacción
            # en la que corre procesar_layout_basica al aceptar la carga).
            try:
                with transaction.atomic():
                    servidor, creado = ServidorPublico.objects.get_or_create(
                        rfc=rfc,
                        defaults={
                            'curp':                  curp,
                            'determinante':          sd(col(C_DETERMINANTE, '')),
                            'expediente':            expediente,
                            'nombre':                nombre,
                            'primer_apellido':       ap_pat,
                            'segundo_apellido':      ap_mat,
                            'fecha_nacimiento':      parse_fecha(col(C_FECHA_NAC)) or date(1900, 1, 1),
                            'sexo':                  normalizar_sexo(col(C_GENERO)),
                            'estado_civil':          cache.estado_civil(col(C_ESTADO_CIVIL)),
                            'entidad_nacimiento':    cache.entidad(col(C_ENTIDAD)),
                            'pais_nacimiento':       cache.pais(col(C_PAIS)),
                            'correo_institucional':  sd(col(C_CORREO, '')),
                            'iss':                   normalizar_iss(col(C_ISS)),
                            'nss':                   sd(col(C_NSS, '')),
                            'sindicalizado':         normalizar_sino(col(C_SINDICALIZADO)),
                            'sindicato':             cache.sindicato(col(C_SINDICATO)),
                            'tiene_otra_plaza':      normalizar_sino(col(C_OTRA_PLAZA)),
                            'activo':                True,
                        }
                    )
                    if not creado:
                        # Actualizar campos que pueden cambiar quincenalmente
                        cambios = False
                        nuevo_correo = sd(col(C_CORREO, ''))
                        nuevo_nss    = sd(col(C_NSS, ''))
                        nueva_determinante = sd(col(C_DETERMINANTE, ''))
                        if nuevo_correo and servidor.correo_institucional != nuevo_correo:
                            servidor.correo_institucional = nuevo_correo; cambios = True
                        if nuevo_nss and servidor.nss != nuevo_nss:
                            servidor.nss = nuevo_nss; cambios = True
                        if nueva_determinante and servidor.determinante != nueva_determinante:
                            servidor.determinante = nueva_determinante; cambios = True
                        if ap_mat and servidor.segundo_apellido != ap_mat:
                            servidor.segundo_apellido = ap_mat; cambios = True
                        ec = cache.estado_civil(col(C_ESTADO_CIVIL))
                        if ec and servidor.estado_civil != ec:
                            servidor.estado_civil = ec; cambios = True
                        sin = cache.sindicato(col(C_SINDICATO))
                        if sin and servidor.sindicato != sin:
                            servidor.sindicato = sin; cambios = True
                        if not servidor.activo:
                            servidor.activo = True; cambios = True; reactivado = True
                        if cambios:
                            servidor.save()

            except Exception as e:
                # Posible CURP duplicado con otro RFC
                try:
                    servidor = ServidorPublico.objects.get(curp=curp)
                except ServidorPublico.DoesNotExist:
                    errores += 1
                    log.append(f'Fila {num_fila} RFC={rfc}: ERROR creando servidor — {e}')
                    filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp,
                                           'nombre': f'{nombre} {ap_pat} {ap_mat or ""}'.strip(), 'id_plaza': id_plaza,
                                           'estado': 'ERROR', 'mensaje': f'Error creando servidor — {e}'})
                    continue

        # ── Catálogos del puesto ─────────────────────────────────────────────
        fuente      = cache.fuente(col(C_FTE_FINAN))
        dependencia = cache.dependencia(col(C_DEPENDENCIA))

        # La dependencia de la fila (si se reconoce en el catálogo) debe ser
        # la misma que se eligió al subir el archivo — no se permite mezclar
        # datos de otra dependencia en la carga de una dependencia distinta.
        if dependencia and dependencia.pk != carga.dependencia_id:
            errores += 1
            msg = (f'La dependencia de la fila ({dependencia.clave}) no coincide con '
                   f'la dependencia de la carga ({carga.dependencia.clave})')
            log.append(f'Fila {num_fila} RFC={rfc}: ERROR — {msg}')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp,
                                   'nombre': f'{nombre} {ap_pat} {ap_mat or ""}'.strip(), 'id_plaza': id_plaza,
                                   'estado': 'ERROR', 'mensaje': msg})
            continue

        unidad      = cache.unidad(col(C_UNIDAD))
        programa    = cache.programa(col(C_PROGRAMA))
        proyecto    = cache.proyecto(col(C_PROYECTO), dependencia=dependencia)
        categoria   = cache.categoria(col(C_CATEGORIA))
        nombramiento= cache.tipo_contratacion(col(C_TIPO_CONTRA))
        tipo_pers   = cache.tipo_personal(col(C_TIPO_PERSONAL))
        tipo_fun    = cache.tipo_funcion(col(C_TIPO_FUNCION))
        nivel_est   = cache.nivel_estructura(col(C_NIVEL_EST))
        estatus     = cache.estatus_plaza(col(C_ESTATUS_PLAZA))
        cct         = cache.centro_trabajo(col(C_CCT))
        tipo_decla  = cache.tipo_declaracion(col(C_TIPO_DECLA))
        area        = cache.area(col(C_AREA))
        inmueble    = cache.inmueble(col(C_INMUEBLE))

        # Resto de columnas del layout que van a InformacionBasica — se calculan
        # una sola vez aquí para poder compararlas contra el período anterior
        # y para no repetir la lectura al crear el registro más abajo.
        id_plaza_jefe_val           = sd(col(C_ID_PUESTO_JEFE, ''))
        hsm_val                     = sf(col(C_HSM))
        percepciones_val            = sf(col(C_PERCEPCIONES))
        bonos_val                   = sf(col(C_BONOS))
        neto_val                    = sf(col(C_NETO))
        dias_pagados_val            = si(col(C_DIAS_PAGADOS))
        oblig_declaracion_val       = normalizar_sino(col(C_ORDP))
        oblig_entrega_recepcion_val = normalizar_sino(col(C_OPAAER))
        oblig_rendir_cuentas_val    = normalizar_sino(col(C_RENCTA))
        puesto_sensible_val         = normalizar_sino(col(C_PRDMIS))
        f_ingob_val                 = parse_fecha(col(C_F_INGOB))
        f_indep_val                 = parse_fecha(col(C_F_INDEP))
        f_inpues_val                = parse_fecha(col(C_F_INPUES))
        participa_contrataciones_val= normalizar_sino(col(C_PARCONPUB))
        nivel_contrataciones_val    = sd(col(C_CONTRAT_PUB, ''))
        participa_concesiones_val   = normalizar_sino(col(C_PARCON))
        nivel_concesiones_val       = sd(col(C_CONCESIONES, ''))
        participa_enajenacion_val   = normalizar_sino(col(C_PARENA))
        nivel_enajenacion_val       = sd(col(C_ENAJENACION, ''))

        avisos = list(avisos_forzados)
        if not dependencia:
            avisos.append(f'Dependencia "{col(C_DEPENDENCIA)}" no encontrada')
        if not estatus:
            avisos.append(f'Estatus plaza "{col(C_ESTATUS_PLAZA)}" no encontrado')
        if not nombramiento:
            avisos.append(f'Tipo contratación "{col(C_TIPO_CONTRA)}" no encontrado')
        if (proyecto and programa and proyecto.clave != '00000000'
                and not proyecto.programas.filter(pk=programa.pk).exists()):
            avisos.append(f'El programa "{programa.clave}" no está asignado al proyecto "{proyecto.clave}"')
        if not curp_digito_verificador_valido(curp):
            avisos.append('Dígito verificador de CURP no coincide (posible error de captura)')
        if reactivado:
            avisos.append('El servidor estaba inactivo (dado de baja) y se reactiva con esta carga')

        # ── Comparación contra el período anterior (aviso, no bloquea) ───────
        valores_nuevos = {
            'fuente_financiamiento':     fuente,
            'unidad':                    unidad,
            'programa':                  programa,
            'proyecto':                  proyecto,
            'categoria':                 categoria,
            'nombramiento':              nombramiento,
            'tipo_personal':             tipo_pers,
            'tipo_funcion':              tipo_fun,
            'nivel_estructura':          nivel_est,
            'id_plaza_jefe':             id_plaza_jefe_val,
            'hsm':                       hsm_val,
            'total_percepciones':        percepciones_val,
            'total_bonos':               bonos_val,
            'total_neto':                neto_val,
            'dias_pagados':              dias_pagados_val,
            'estatus_plaza':             estatus,
            'cct':                       cct,
            'oblig_declaracion':         oblig_declaracion_val,
            'tipo_declaracion':          tipo_decla,
            'oblig_entrega_recepcion':   oblig_entrega_recepcion_val,
            'oblig_rendir_cuentas':      oblig_rendir_cuentas_val,
            'puesto_sensible':           puesto_sensible_val,
            'fecha_ingreso_gobierno':    f_ingob_val,
            'fecha_ingreso_dependencia': f_indep_val,
            'fecha_ingreso_puesto':      f_inpues_val,
            'area':                      area,
            'participa_contrataciones':  participa_contrataciones_val,
            'nivel_contrataciones':      nivel_contrataciones_val,
            'participa_concesiones':     participa_concesiones_val,
            'nivel_concesiones':         nivel_concesiones_val,
            'participa_enajenacion':     participa_enajenacion_val,
            'nivel_enajenacion':         nivel_enajenacion_val,
            'inmueble':                  inmueble,
        }
        cambios_periodo = comparar_con_periodo_anterior(servidor, id_plaza, quincena, valores_nuevos)
        if cambios_periodo:
            avisos.append(f'Cambios vs. período anterior ({len(cambios_periodo)}): ' + '; '.join(cambios_periodo))

        nombre_completo = f'{nombre} {ap_pat} {ap_mat or ""}'.strip()

        if dry_run:
            # Vista previa: no se desactiva el registro anterior ni se crea el
            # nuevo; solo se cuenta como OK con los mismos avisos que tendría
            # la aplicación real.
            ok += 1
            accion = 'NUEVO' if creado else 'ACTUALIZADO'
            if avisos:
                log.append(f'Fila {num_fila} RFC={rfc}: OK ({accion}) — Avisos: {"; ".join(avisos)}')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp, 'nombre': nombre_completo,
                                   'id_plaza': id_plaza, 'estado': 'OK',
                                   'mensaje': '; '.join(avisos)})
            continue

        # ── Desactivar registro anterior + crear Información Básica ──────────
        # "activo" debe representar la situación VIGENTE de este servidor en
        # esta plaza — así que se desactiva cualquier InformacionBasica activa
        # anterior de ese mismo servidor+plaza, sea de esta misma quincena
        # (corrección de una carga ya aplicada) o de una quincena anterior
        # (avance normal de quincena a quincena). Sin el filtro de quincena,
        # antes se quedaban "Vigente" varias quincenas a la vez para la misma
        # plaza. Se acota a servidor+id_plaza (no solo servidor) para no tocar
        # otras plazas que ese mismo servidor tenga por compatibilidad.
        # En un solo savepoint: si esta fila falla, se revierte completa (no
        # deja el puesto anterior desactivado sin el registro nuevo) y no
        # invalida la transacción en la que corren el resto de las filas.
        try:
            with transaction.atomic():
                InformacionBasica.objects.filter(
                    servidor=servidor,
                    id_plaza=id_plaza,
                    activo=True,
                ).update(activo=False)

                ib_creada = InformacionBasica.objects.create(
                    carga_origen               = carga,
                    # Institución
                    fuente_financiamiento      = fuente,
                    dependencia                = dependencia,
                    unidad                     = unidad,
                    programa                   = programa,
                    proyecto                   = proyecto,
                    # Puesto
                    id_plaza                   = id_plaza,
                    categoria                  = categoria,
                    puesto                     = id_plaza,
                    nombramiento               = nombramiento,
                    tipo_personal              = tipo_pers,
                    tipo_funcion               = tipo_fun,
                    nivel_estructura           = nivel_est,
                    id_plaza_jefe              = id_plaza_jefe_val,
                    puesto_jefe                = id_plaza_jefe_val,
                    hsm                        = hsm_val,
                    total_percepciones         = percepciones_val,
                    total_bonos                = bonos_val,
                    total_neto                 = neto_val,
                    dias_pagados               = dias_pagados_val,
                    estatus_plaza              = estatus,
                    # Servidor
                    servidor                   = servidor,
                    cct                        = cct,
                    # Persona-Puesto
                    oblig_declaracion          = oblig_declaracion_val,
                    tipo_declaracion           = tipo_decla,
                    oblig_entrega_recepcion    = oblig_entrega_recepcion_val,
                    oblig_rendir_cuentas       = oblig_rendir_cuentas_val,
                    puesto_sensible            = puesto_sensible_val,
                    # Fechas
                    fecha_ingreso_gobierno     = f_ingob_val,
                    fecha_ingreso_dependencia  = f_indep_val,
                    fecha_ingreso_puesto       = f_inpues_val,
                    # Responsabilidades
                    area                       = area,
                    participa_contrataciones   = participa_contrataciones_val,
                    nivel_contrataciones       = nivel_contrataciones_val,
                    participa_concesiones      = participa_concesiones_val,
                    nivel_concesiones          = nivel_concesiones_val,
                    participa_enajenacion      = participa_enajenacion_val,
                    nivel_enajenacion          = nivel_enajenacion_val,
                    participa_avaluos          = 'N',
                    nivel_avaluos              = '',
                    # Inmueble
                    inmueble                   = inmueble,
                    serc                       = 'NULL',
                    exigibilidad_serc          = 'NULL',
                    # Control
                    quincena                   = quincena,
                    activo                     = True,
                )
                sincronizar_puesto(
                    proyecto, programa, id_plaza, categoria, servidor,
                    unidad=unidad,
                    nombramiento=nombramiento,
                    nivel_estructura=nivel_est,
                    estatus_plaza=estatus,
                    cct=cct,
                    hsm=hsm_val,
                    total_percepciones=percepciones_val,
                    total_bonos=bonos_val,
                    total_neto=neto_val,
                    dias_pagados=dias_pagados_val,
                    id_plaza_jefe=id_plaza_jefe_val,
                )

            ok += 1
            accion = 'NUEVO' if creado else 'ACTUALIZADO'
            if avisos:
                log.append(f'Fila {num_fila} RFC={rfc}: OK ({accion}) — Avisos: {"; ".join(avisos)}')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp, 'nombre': nombre_completo,
                                   'id_plaza': id_plaza, 'estado': 'OK',
                                   'mensaje': '; '.join(avisos),
                                   'informacion_basica_id': ib_creada.pk})

        except Exception as e:
            errores += 1
            log.append(f'Fila {num_fila} RFC={rfc}: ERROR al crear registro — {e}')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp, 'nombre': nombre_completo,
                                   'id_plaza': id_plaza, 'estado': 'ERROR',
                                   'mensaje': f'Error al crear registro — {e}'})

    return {
        'ok':      ok,
        'errores': errores,
        'total':   total,
        'log':     '\n'.join(log),
        'filas':   filas_detalle,
    }


# ── Índices de columna del Layout de Bajas (base 0) ────────────────────────────
CB_DEPENDENCIA  = 0
CB_EJERCICIO    = 1
CB_PERIODO      = 2
CB_RFC          = 3
CB_CURP         = 4
CB_PATERNO      = 5
CB_MATERNO      = 6
CB_NOMBRE       = 7
CB_FECHA_BAJA   = 8
CB_MOTIVO       = 9
CB_EXPEDIENTE   = 10
CB_DETERMINANTE = 11
CB_ID_PLAZA     = 12


def procesar_layout_bajas(carga, dry_run=True, overrides=None):
    """Igual que procesar_layout_basica: dry_run=True valida sin escribir
    (vista previa al subir), dry_run=False aplica de verdad (al aceptar).
    'overrides' funciona igual que en procesar_layout_basica (ver ahí).

    A diferencia del alta manual de baja (servidores/views.py, que libera
    TODAS las plazas del servidor con liberar_puestos_de), aquí la plaza a
    liberar viene explícita en la columna ID_PLAZA de cada fila y se libera
    ÚNICAMENTE esa — necesario porque un servidor puede tener más de una
    plaza (compatibilidad) y la baja de una no implica la de las demás.
    El servidor y la plaza deben existir de antes: este layout no crea
    servidores ni plazas, solo registra su baja/liberación."""
    overrides = overrides or {}
    ruta = carga.archivo.path
    log = []
    filas_detalle = []
    ok = errores = total = 0

    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
    except Exception as e:
        return {'ok': 0, 'errores': 1, 'total': 0, 'log': f'No se pudo abrir el archivo: {e}', 'filas': []}

    hoja = wb['Layout_Bajas'] if 'Layout_Bajas' in wb.sheetnames else wb.active
    filas = list(hoja.iter_rows(values_only=True))

    # Los encabezados de este layout traen subtítulo en la misma celda
    # ("RFC\nRFC con homoclave"), por eso la búsqueda es por substring y no
    # por coincidencia exacta como en procesar_layout_basica.
    inicio_datos = 1
    for idx, fila in enumerate(filas):
        fila_str = [str(c).upper().strip() if c else '' for c in fila]
        if any('RFC' in s for s in fila_str) or any('CURP' in s for s in fila_str):
            inicio_datos = idx + 1
            break

    cache = Cache()

    for num_fila, fila in enumerate(filas[inicio_datos:], start=inicio_datos + 1):
        if not any(c is not None for c in fila):
            continue

        total += 1

        def col(i, default=None):
            try:
                v = fila[i]
                if v is None or str(v).strip().upper() in ('NULL', 'NONE', 'N/A', ''):
                    return default
                return v
            except IndexError:
                return default

        rfc         = sv(col(CB_RFC))
        curp        = sv(col(CB_CURP))
        paterno     = sd(col(CB_PATERNO, '')).upper()
        nombre      = sd(col(CB_NOMBRE, '')).upper()
        id_plaza    = sd(col(CB_ID_PLAZA, ''))
        motivo_clave= sv(col(CB_MOTIVO))
        fecha_baja  = parse_fecha(col(CB_FECHA_BAJA))
        nombre_completo = f'{nombre} {paterno}'.strip()

        errores_fila = []
        errores_fila_forzables = []
        if not rfc:
            errores_fila.append('RFC vacío')
        elif not rfc_formato_valido(rfc):
            errores_fila_forzables.append('RFC con formato inválido')
        if not curp:
            errores_fila.append('CURP vacío')
        elif not curp_formato_valido(curp):
            errores_fila_forzables.append('CURP con formato inválido')
        if not fecha_baja:
            errores_fila.append('Fecha de baja vacía o con formato inválido')
        if not motivo_clave:
            errores_fila.append('Motivo de baja vacío')
        if not id_plaza:
            errores_fila.append('ID de plaza vacío')
        elif not Puesto.objects.filter(id_plaza=id_plaza).exists():
            errores_fila.append(f'La plaza {id_plaza} no existe (debe darse de alta primero en la carga inicial de estructura)')

        avisos_forzados = []
        override = overrides.get(num_fila)
        if errores_fila_forzables:
            if not errores_fila and override and override.get('decision') == 'aceptada':
                motivo = override.get('motivo', '')
                avisos_forzados = [f'{msg} — ACEPTADO por el validador: {motivo}' for msg in errores_fila_forzables]
            else:
                errores_fila.extend(errores_fila_forzables)

        if errores_fila:
            errores += 1
            log.append(f'Fila {num_fila}: OMITIDA — {", ".join(errores_fila)}')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp, 'nombre': nombre_completo,
                                   'id_plaza': id_plaza, 'estado': 'OMITIDA', 'mensaje': ', '.join(errores_fila),
                                   'forzable': es_fila_forzable(errores_fila)})
            continue

        servidor = ServidorPublico.objects.filter(rfc=rfc).first()
        if not servidor:
            errores += 1
            log.append(f'Fila {num_fila} RFC={rfc}: ERROR — no existe un servidor con ese RFC en el padrón')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp, 'nombre': nombre_completo,
                                   'id_plaza': id_plaza, 'estado': 'ERROR',
                                   'mensaje': 'No existe un servidor con ese RFC en el padrón'})
            continue

        motivo = cache.motivo_baja(motivo_clave)
        if not motivo:
            errores += 1
            log.append(f'Fila {num_fila} RFC={rfc}: ERROR — motivo de baja "{motivo_clave}" no encontrado en el catálogo')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp, 'nombre': nombre_completo,
                                   'id_plaza': id_plaza, 'estado': 'ERROR',
                                   'mensaje': f'Motivo de baja "{motivo_clave}" no encontrado en el catálogo'})
            continue

        dependencia    = cache.dependencia(col(CB_DEPENDENCIA))
        if dependencia and dependencia.pk != carga.dependencia_id:
            errores += 1
            msg = (f'La dependencia de la fila ({dependencia.clave}) no coincide con '
                   f'la dependencia de la carga ({carga.dependencia.clave})')
            log.append(f'Fila {num_fila} RFC={rfc}: ERROR — {msg}')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp, 'nombre': nombre_completo,
                                   'id_plaza': id_plaza, 'estado': 'ERROR', 'mensaje': msg})
            continue
        ejercicio_col  = si(col(CB_EJERCICIO), 0)
        periodo_col    = sd(col(CB_PERIODO, ''))

        avisos = list(avisos_forzados)
        if not dependencia:
            avisos.append(f'Dependencia "{col(CB_DEPENDENCIA)}" no encontrada')
        puesto_actual = Puesto.objects.filter(id_plaza=id_plaza).select_related('servidor_actual').first()
        if puesto_actual and puesto_actual.servidor_actual_id != servidor.pk:
            avisos.append(f'La plaza {id_plaza} no estaba asignada a este servidor al momento de la baja')

        if dry_run:
            ok += 1
            if avisos:
                log.append(f'Fila {num_fila} RFC={rfc}: OK — Avisos: {"; ".join(avisos)}')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp, 'nombre': nombre_completo,
                                   'id_plaza': id_plaza, 'estado': 'OK', 'mensaje': '; '.join(avisos)})
            continue

        # Fila completa en un solo savepoint: registrar la baja + liberar
        # exactamente esa plaza + desactivar solo el InformacionBasica de esa
        # plaza + dejar inactivo al servidor únicamente si ya no le queda
        # ninguna otra plaza asignada (puede tener más de una por compatibilidad).
        try:
            with transaction.atomic():
                baja_creada = BajaServidorPublico.objects.create(
                    servidor=servidor,
                    dependencia=dependencia or carga.dependencia,
                    id_plaza=id_plaza,
                    fecha_baja=fecha_baja,
                    motivo_baja=motivo,
                    ejercicio=ejercicio_col or carga.periodo.ejercicio,
                    periodo=periodo_col or f'{carga.periodo.ejercicio}-{carga.periodo.quincena}',
                    registrado_por=carga.usuario_carga,
                    carga_origen=carga,
                )
                reportar_puesto_vacante(id_plaza)
                InformacionBasica.objects.filter(
                    servidor=servidor, id_plaza=id_plaza, activo=True
                ).update(activo=False)
                if not Puesto.objects.filter(servidor_actual=servidor).exists():
                    servidor.activo = False
                    servidor.save()

            ok += 1
            if avisos:
                log.append(f'Fila {num_fila} RFC={rfc}: OK — Avisos: {"; ".join(avisos)}')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp, 'nombre': nombre_completo,
                                   'id_plaza': id_plaza, 'estado': 'OK', 'mensaje': '; '.join(avisos),
                                   'baja_id': baja_creada.pk})

        except Exception as e:
            errores += 1
            log.append(f'Fila {num_fila} RFC={rfc}: ERROR al registrar la baja — {e}')
            filas_detalle.append({'fila': num_fila, 'rfc': rfc, 'curp': curp, 'nombre': nombre_completo,
                                   'id_plaza': id_plaza, 'estado': 'ERROR',
                                   'mensaje': f'Error al registrar la baja — {e}'})

    return {
        'ok':      ok,
        'errores': errores,
        'total':   total,
        'log':     '\n'.join(log),
        'filas':   filas_detalle,
    }
