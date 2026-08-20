# -*- coding: utf-8 -*-
import os
import threading
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.http import FileResponse, HttpResponse, Http404
from django.db.models import Q, ProtectedError
from django.db.models.deletion import Collector
from django.core.files.storage import default_storage
from django.contrib import messages

from usuarios.mixins import AdministradorRequiredMixin, DependenciaScopedMixin, DependenciaFormRestrictMixin, filtrar_por_dependencia, admin_requerido

from .models import (
    FuenteFinanciamiento, Dependencia, UnidadAdministrativa,
    Programa, Proyecto, Categoria, TipoContratacion, TipoPersonal,
    TipoFuncion, NivelEstructura, EstatusPlaza, CentroTrabajo,
    TipoDeclaracion, Area, NivelEscolaridad, Discapacidad,
    EnfermedadCronica, Pueblo, MotivoBaja, Idioma, EstadoCivil,
    Pais, EntidadFederativa, Municipio, Sindicato, Inmueble,
    ImportacionCatalogo,
)
from .forms import (
    FuenteFinanciamientoForm, DependenciaForm, UnidadAdministrativaForm,
    ProgramaForm, ProyectoForm, CategoriaForm, TipoContratacionForm,
    TipoPersonalForm, TipoFuncionForm, NivelEstructuraForm, EstatusPlazaForm,
    CentroTrabajoForm, TipoDeclaracionForm, AreaForm, NivelEscolaridadForm,
    DiscapacidadForm, EnfermedadCronicaForm, PuebloForm, MotivoBajaForm,
    IdiomaForm, EstadoCivilForm, PaisForm, EntidadFederativaForm,
    MunicipioForm, SindicatoForm, InmuebleForm, ImportarExcelForm,
)

# ── Índice ────────────────────────────────────────────────────────────────────
def catalogo_index(request):
    grupos = [
        {
            'nombre': 'Estructura Presupuestal',
            'color':  'verde',
            'items': [
                {'nombre': 'Fuentes de Financiamiento', 'url': 'fuente_list',    'icono': '💰', 'descarga': 'fuente'},
                {'nombre': 'Dependencias',               'url': 'dependencia_list','icono': '🏛️', 'descarga': 'dependencia'},
                {'nombre': 'Unidades Administrativas',   'url': 'unidad_list',   'icono': '🏢', 'descarga': 'unidades'},
                {'nombre': 'Programas',                  'url': 'programa_list', 'icono': '📋', 'descarga': 'programas'},
                {'nombre': 'Proyectos',                  'url': 'proyecto_list', 'icono': '📁', 'descarga': 'proyectos'},
            ]
        },
        {
            'nombre': 'Plazas',
            'color':  'azul',
            'items': [
                {'nombre': 'Plazas',              'url': 'puesto_list',          'icono': '📌',  'descarga': 'puestos'},
                {'nombre': 'Categorías',          'url': 'categoria_list',       'icono': '🏷️',  'descarga': 'categoria'},
                {'nombre': 'Tipos de Contratación','url': 'tipo_contratacion_list','icono': '📝', 'descarga': None},
                {'nombre': 'Tipos de Personal',   'url': 'tipo_personal_list',   'icono': '👤',  'descarga': None},
                {'nombre': 'Tipos de Función',    'url': 'tipo_funcion_list',    'icono': '⚙️',  'descarga': None},
                {'nombre': 'Niveles de Estructura','url': 'nivel_estructura_list','icono': '🔢',  'descarga': None},
                {'nombre': 'Estatus de Plaza',    'url': 'estatus_plaza_list',   'icono': '📊',  'descarga': None},
                {'nombre': 'Centros de Trabajo',  'url': 'centro_trabajo_list',  'icono': '🏫',  'descarga': None},
            ]
        },
        {
            'nombre': 'Datos del Servidor Público',
            'color':  'dorado',
            'items': [
                {'nombre': 'Estados Civiles',         'url': 'estado_civil_list',     'icono': '💍', 'descarga': None},
                {'nombre': 'Entidades Federativas',   'url': 'entidad_list',          'icono': '🗺️', 'descarga': None},
                {'nombre': 'Municipios',              'url': 'municipio_list',        'icono': '📍', 'descarga': None},
                {'nombre': 'Países',                  'url': 'pais_list',             'icono': '🌎', 'descarga': None},
                {'nombre': 'Sindicatos',              'url': 'sindicato_list',        'icono': '🤝', 'descarga': None},
                {'nombre': 'Niveles de Escolaridad',  'url': 'nivel_escolaridad_list','icono': '🎓', 'descarga': None},
                {'nombre': 'Idiomas / Lenguas',       'url': 'idioma_list',           'icono': '🗣️', 'descarga': None},
                {'nombre': 'Discapacidades',          'url': 'discapacidad_list',     'icono': '♿', 'descarga': None},
                {'nombre': 'Enfermedades Crónicas',   'url': 'enfermedad_list',       'icono': '🏥', 'descarga': None},
                {'nombre': 'Pueblos Indígenas',       'url': 'pueblo_list',           'icono': '🪶', 'descarga': None},
            ]
        },
        {
            'nombre': 'Otros Catálogos',
            'color':  'rojo',
            'items': [
                {'nombre': 'Tipos de Declaración',  'url': 'tipo_declaracion_list', 'icono': '⚖️', 'descarga': None},
                {'nombre': 'Áreas',                 'url': 'area_list',             'icono': '🏗️', 'descarga': None},
                {'nombre': 'Motivos de Baja',       'url': 'motivo_baja_list',      'icono': '🚫', 'descarga': None},
                {'nombre': 'Inmuebles',             'url': 'inmueble_list',         'icono': '🏠', 'descarga': None},
            ]
        },
    ]
    from .importador import IMPORTADORES
    for grupo in grupos:
        for item in grupo['items']:
            item['importar'] = item['descarga'] in IMPORTADORES and request.user.es_administrador

    return render(request, 'catalogos/index.html', {
        'titulo': 'Catálogos del Sistema',
        'grupos': grupos,
    })


# ── Mixin reutilizable ────────────────────────────────────────────────────────
class CatalogoMixin(LoginRequiredMixin):
    """Mixin base para vistas de catálogos: agrega titulo y back_url al contexto."""
    titulo       = ''
    back_url_name = ''

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo']   = self.titulo
        ctx['back_url'] = reverse_lazy(self.back_url_name) if self.back_url_name else '#'
        return ctx


def make_list_view(model, template, titulo, search_fields=None, paginate=30, extra_ctx=None, dependencia_lookup=None):
    """Factoría de ListViews para catálogos simples. 'dependencia_lookup' activa
    el filtrado por dependencia del usuario (None = catálogo global, sin filtrar)."""
    class V(LoginRequiredMixin, ListView):
        queryset         = model.objects.all()
        template_name    = template
        context_object_name = 'registros'
        paginate_by      = paginate

        def get_queryset(self):
            qs = super().get_queryset()
            if dependencia_lookup:
                qs = filtrar_por_dependencia(qs, self.request.user, dependencia_lookup)
            q  = self.request.GET.get('q', '')
            if q and search_fields:
                filtro = Q()
                for f in search_fields:
                    filtro |= Q(**{f'{f}__icontains': q})
                qs = qs.filter(filtro)
            return qs

        def get_context_data(self, **kwargs):
            ctx = super().get_context_data(**kwargs)
            ctx['titulo'] = titulo
            ctx['q']      = self.request.GET.get('q', '')
            if extra_ctx:
                ctx.update(extra_ctx() if callable(extra_ctx) else extra_ctx)
            return ctx
    return V


def make_create_view(model, form_class, success_url, titulo, back_url, template='catalogos/form_generica.html', mixins=()):
    """'mixins' se insertan antes de CreateView (ej. DependenciaFormRestrictMixin)."""
    class V(*mixins, CatalogoMixin, CreateView):
        pass
    V.model        = model
    V.form_class   = form_class
    V.template_name      = template
    V.success_url        = reverse_lazy(success_url)
    V.titulo             = titulo
    V.back_url_name      = back_url
    return V


def make_update_view(model, form_class, success_url, titulo_prefix, back_url, template='catalogos/form_generica.html', mixins=()):
    """'mixins' se insertan antes de UpdateView (ej. DependenciaFormRestrictMixin)."""
    class V(*mixins, LoginRequiredMixin, UpdateView):
        def get_context_data(self, **kwargs):
            ctx = super().get_context_data(**kwargs)
            ctx['titulo']   = f'{titulo_prefix}: {self.object}'
            ctx['back_url'] = reverse_lazy(back_url)
            return ctx
    V.model        = model
    V.form_class   = form_class
    V.template_name      = template
    V.success_url        = reverse_lazy(success_url)
    return V


def make_delete_view(model, success_url, titulo_prefix, back_url, template='catalogos/confirm_delete.html'):
    """Factoría de vistas de eliminación con confirmación. Muestra qué otros
    registros se borrarían en cascada y evita romper la app si el catálogo
    está protegido por alguna relación (PROTECT)."""
    class V(LoginRequiredMixin, DeleteView):
        def get_context_data(self, **kwargs):
            ctx = super().get_context_data(**kwargs)
            ctx['titulo']   = f'{titulo_prefix}: {self.object}'
            ctx['back_url'] = reverse_lazy(back_url)
            collector = Collector(using=self.object._state.db)
            collector.collect([self.object])
            ctx['relacionados'] = [
                (rel_model._meta.verbose_name_plural, len(instancias))
                for rel_model, instancias in collector.data.items()
                if rel_model is not model
            ]
            return ctx

        def form_valid(self, form):
            descripcion = str(self.object)
            try:
                response = super().form_valid(form)
            except ProtectedError:
                messages.error(
                    self.request,
                    f'No se puede eliminar "{descripcion}" porque está en uso en otros registros.'
                )
                return redirect(reverse_lazy(success_url))
            messages.success(self.request, f'"{descripcion}" fue eliminado.')
            return response
    V.model        = model
    V.template_name = template
    V.success_url   = reverse_lazy(success_url)
    return V


# ── Fuente de Financiamiento ──────────────────────────────────────────────────
FuenteListView   = make_list_view(FuenteFinanciamiento, 'catalogos/simple_list.html',
                                  'Fuentes de Financiamiento', ['clave', 'descripcion'],
                                  extra_ctx={'catalogo_slug': 'fuente', 'export_slug': 'fuente'})
FuenteCreateView = make_create_view(FuenteFinanciamiento, FuenteFinanciamientoForm,
                                    'fuente_list', 'Nueva Fuente de Financiamiento', 'fuente_list')
FuenteUpdateView = make_update_view(FuenteFinanciamiento, FuenteFinanciamientoForm,
                                    'fuente_list', 'Editar Fuente', 'fuente_list')

# ── Dependencias ──────────────────────────────────────────────────────────────
# Alta de nuevas dependencias (agencias de gobierno) es una acción administrativa;
# un usuario no-admin solo puede consultar (y solo ve) la suya propia.
DependenciaListView   = make_list_view(Dependencia, 'catalogos/simple_list.html',
                                       'Dependencias', ['clave', 'descripcion'],
                                       dependencia_lookup='pk',
                                       extra_ctx={'catalogo_slug': 'dependencia', 'export_slug': 'dependencia'})
DependenciaCreateView = make_create_view(Dependencia, DependenciaForm,
                                         'dependencia_list', 'Nueva Dependencia', 'dependencia_list',
                                         mixins=(AdministradorRequiredMixin,))
DependenciaUpdateView = make_update_view(Dependencia, DependenciaForm,
                                         'dependencia_list', 'Editar Dependencia', 'dependencia_list',
                                         mixins=(AdministradorRequiredMixin,))

# ── Unidades Administrativas ──────────────────────────────────────────────────
class UnidadListView(LoginRequiredMixin, ListView):
    model               = UnidadAdministrativa
    template_name       = 'catalogos/unidad_list.html'
    context_object_name = 'registros'
    paginate_by         = 40

    def get_queryset(self):
        qs  = filtrar_por_dependencia(
            UnidadAdministrativa.objects.select_related('dependencia'), self.request.user
        )
        q   = self.request.GET.get('q', '')
        dep = self.request.GET.get('dep', '')
        if q:
            qs = qs.filter(Q(clave__icontains=q) | Q(descripcion__icontains=q))
        if dep:
            qs = qs.filter(dependencia_id=dep)
        return qs.order_by('dependencia__clave', 'clave')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({'titulo': 'Unidades Administrativas',
                    'dependencias': filtrar_por_dependencia(Dependencia.objects.all(), self.request.user, 'pk').order_by('clave'),
                    'q': self.request.GET.get('q', ''),
                    'dep_sel': self.request.GET.get('dep', ''),
                    'export_slug': 'unidad'})
        return ctx

UnidadCreateView = make_create_view(UnidadAdministrativa, UnidadAdministrativaForm,
                                    'unidad_list', 'Nueva Unidad Administrativa', 'unidad_list',
                                    mixins=(DependenciaFormRestrictMixin,))
UnidadUpdateView = make_update_view(UnidadAdministrativa, UnidadAdministrativaForm,
                                    'unidad_list', 'Editar Unidad', 'unidad_list',
                                    mixins=(DependenciaFormRestrictMixin, DependenciaScopedMixin))

# ── Programas ─────────────────────────────────────────────────────────────────
class ProgramaListView(LoginRequiredMixin, ListView):
    model               = Programa
    template_name       = 'catalogos/programa_list.html'
    context_object_name = 'registros'
    paginate_by         = 40

    def get_queryset(self):
        qs  = filtrar_por_dependencia(Programa.objects.select_related('dependencia', 'unidad'), self.request.user)
        q   = self.request.GET.get('q', '')
        dep = self.request.GET.get('dep', '')
        if q:
            qs = qs.filter(Q(clave__icontains=q) | Q(descripcion__icontains=q))
        if dep:
            qs = qs.filter(dependencia_id=dep)
        return qs.order_by('dependencia__clave', 'clave')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({'titulo': 'Programas Presupuestales',
                    'dependencias': filtrar_por_dependencia(Dependencia.objects.all(), self.request.user, 'pk').order_by('clave'),
                    'q': self.request.GET.get('q', ''),
                    'dep_sel': self.request.GET.get('dep', ''),
                    'export_slug': 'programa'})
        return ctx

ProgramaCreateView = make_create_view(Programa, ProgramaForm,
                                      'programa_list', 'Nuevo Programa', 'programa_list',
                                      mixins=(DependenciaFormRestrictMixin,))
ProgramaUpdateView = make_update_view(Programa, ProgramaForm,
                                      'programa_list', 'Editar Programa', 'programa_list',
                                      mixins=(DependenciaFormRestrictMixin, DependenciaScopedMixin))

# ── Proyectos ─────────────────────────────────────────────────────────────────
class ProyectoListView(LoginRequiredMixin, ListView):
    model               = Proyecto
    template_name       = 'catalogos/proyecto_list.html'
    context_object_name = 'registros'
    paginate_by         = 40

    def get_queryset(self):
        qs  = filtrar_por_dependencia(
            Proyecto.objects.select_related('dependencia').prefetch_related('programas__unidad'), self.request.user
        )
        q   = self.request.GET.get('q', '')
        dep = self.request.GET.get('dep', '')
        if q:
            qs = qs.filter(Q(clave__icontains=q) | Q(descripcion__icontains=q))
        if dep:
            qs = qs.filter(dependencia_id=dep)
        return qs.order_by('dependencia__clave', 'clave').distinct()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({'titulo': 'Proyectos',
                    'dependencias': filtrar_por_dependencia(Dependencia.objects.all(), self.request.user, 'pk').order_by('clave'),
                    'q': self.request.GET.get('q', ''),
                    'dep_sel': self.request.GET.get('dep', ''),
                    'export_slug': 'proyecto'})
        return ctx

ProyectoCreateView = make_create_view(Proyecto, ProyectoForm,
                                      'proyecto_list', 'Nuevo Proyecto', 'proyecto_list',
                                      template='catalogos/proyecto_form.html',
                                      mixins=(DependenciaFormRestrictMixin,))
ProyectoUpdateView = make_update_view(Proyecto, ProyectoForm,
                                      'proyecto_list', 'Editar Proyecto', 'proyecto_list',
                                      template='catalogos/proyecto_form.html',
                                      mixins=(DependenciaFormRestrictMixin, DependenciaScopedMixin))

# ── Categorías ────────────────────────────────────────────────────────────────
class CategoriaListView(LoginRequiredMixin, ListView):
    model               = Categoria
    template_name       = 'catalogos/categoria_list.html'
    context_object_name = 'registros'
    paginate_by         = 40

    def get_queryset(self):
        qs = Categoria.objects.all()
        q  = self.request.GET.get('q', '')
        if q:
            qs = qs.filter(Q(clave__icontains=q) | Q(descripcion__icontains=q))
        return qs.order_by('nivel', 'clave')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({'titulo': 'Categorías', 'q': self.request.GET.get('q', ''), 'export_slug': 'categoria'})
        return ctx

CategoriaCreateView = make_create_view(Categoria, CategoriaForm,
                                       'categoria_list', 'Nueva Categoría', 'categoria_list')
CategoriaUpdateView = make_update_view(Categoria, CategoriaForm,
                                       'categoria_list', 'Editar Categoría', 'categoria_list')

# ── Catálogos simples (clave + descripcion) ───────────────────────────────────
TipoContratacionListView   = make_list_view(TipoContratacion,  'catalogos/simple_list.html', 'Tipos de Contratación',  ['clave','descripcion'],
                                            extra_ctx={'export_slug': 'tipo_contratacion'})
TipoContratacionCreateView = make_create_view(TipoContratacion,  TipoContratacionForm,  'tipo_contratacion_list', 'Nuevo Tipo Contratación',  'tipo_contratacion_list')
TipoContratacionUpdateView = make_update_view(TipoContratacion,  TipoContratacionForm,  'tipo_contratacion_list', 'Editar',                    'tipo_contratacion_list')

TipoPersonalListView   = make_list_view(TipoPersonal,    'catalogos/simple_list.html', 'Tipos de Personal',      ['clave','descripcion'],
                                        extra_ctx={'export_slug': 'tipo_personal'})
TipoPersonalCreateView = make_create_view(TipoPersonal,    TipoPersonalForm,    'tipo_personal_list',    'Nuevo Tipo Personal',       'tipo_personal_list')
TipoPersonalUpdateView = make_update_view(TipoPersonal,    TipoPersonalForm,    'tipo_personal_list',    'Editar',                    'tipo_personal_list')

TipoFuncionListView   = make_list_view(TipoFuncion,     'catalogos/simple_list.html', 'Tipos de Función',       ['clave','descripcion'],
                                       extra_ctx={'export_slug': 'tipo_funcion'})
TipoFuncionCreateView = make_create_view(TipoFuncion,     TipoFuncionForm,     'tipo_funcion_list',     'Nuevo Tipo Función',         'tipo_funcion_list')
TipoFuncionUpdateView = make_update_view(TipoFuncion,     TipoFuncionForm,     'tipo_funcion_list',     'Editar',                    'tipo_funcion_list')

NivelEstructuraListView   = make_list_view(NivelEstructura,  'catalogos/simple_list.html', 'Niveles de Estructura',  ['clave','descripcion'],
                                           extra_ctx={'export_slug': 'nivel_estructura'})
NivelEstructuraCreateView = make_create_view(NivelEstructura,  NivelEstructuraForm,  'nivel_estructura_list', 'Nuevo Nivel',               'nivel_estructura_list')
NivelEstructuraUpdateView = make_update_view(NivelEstructura,  NivelEstructuraForm,  'nivel_estructura_list', 'Editar',                    'nivel_estructura_list')

EstatusPlazaListView   = make_list_view(EstatusPlaza,    'catalogos/simple_list.html', 'Estatus de Plaza',       ['clave','descripcion'],
                                        extra_ctx={'export_slug': 'estatus_plaza'})
EstatusPlazaCreateView = make_create_view(EstatusPlaza,    EstatusPlazaForm,    'estatus_plaza_list',    'Nuevo Estatus',              'estatus_plaza_list')
EstatusPlazaUpdateView = make_update_view(EstatusPlaza,    EstatusPlazaForm,    'estatus_plaza_list',    'Editar',                    'estatus_plaza_list')

CentroTrabajoListView   = make_list_view(CentroTrabajo,   'catalogos/simple_list.html', 'Centros de Trabajo',     ['clave','nombre'],
                                         extra_ctx={'export_slug': 'centro_trabajo'})
CentroTrabajoCreateView = make_create_view(CentroTrabajo,   CentroTrabajoForm,   'centro_trabajo_list',   'Nuevo Centro de Trabajo',   'centro_trabajo_list')
CentroTrabajoUpdateView = make_update_view(CentroTrabajo,   CentroTrabajoForm,   'centro_trabajo_list',   'Editar',                    'centro_trabajo_list')

TipoDeclaracionListView   = make_list_view(TipoDeclaracion, 'catalogos/simple_list.html', 'Tipos de Declaración',  ['clave','descripcion'],
                                           extra_ctx={'export_slug': 'tipo_declaracion'})
TipoDeclaracionCreateView = make_create_view(TipoDeclaracion, TipoDeclaracionForm, 'tipo_declaracion_list', 'Nuevo Tipo Declaración',    'tipo_declaracion_list')
TipoDeclaracionUpdateView = make_update_view(TipoDeclaracion, TipoDeclaracionForm, 'tipo_declaracion_list', 'Editar',                    'tipo_declaracion_list')

AreaListView   = make_list_view(Area,          'catalogos/simple_list.html', 'Áreas',                  ['clave','descripcion'],
                                extra_ctx={'export_slug': 'area'})
AreaCreateView = make_create_view(Area,          AreaForm,          'area_list',             'Nueva Área',                'area_list')
AreaUpdateView = make_update_view(Area,          AreaForm,          'area_list',             'Editar',                    'area_list')

NivelEscolaridadListView   = make_list_view(NivelEscolaridad, 'catalogos/simple_list.html', 'Niveles de Escolaridad', ['descripcion','estatus'],
                                            extra_ctx={'permite_eliminar': True, 'export_slug': 'nivel_escolaridad'})
NivelEscolaridadCreateView = make_create_view(NivelEscolaridad, NivelEscolaridadForm, 'nivel_escolaridad_list', 'Nuevo Nivel Escolaridad', 'nivel_escolaridad_list')
NivelEscolaridadUpdateView = make_update_view(NivelEscolaridad, NivelEscolaridadForm, 'nivel_escolaridad_list', 'Editar',                  'nivel_escolaridad_list')
NivelEscolaridadDeleteView = make_delete_view(NivelEscolaridad, 'nivel_escolaridad_list', 'Eliminar Nivel de Escolaridad', 'nivel_escolaridad_list')

DiscapacidadListView   = make_list_view(Discapacidad,   'catalogos/simple_list.html', 'Discapacidades',         ['tipo','descripcion'],
                                        extra_ctx={'permite_eliminar': True, 'export_slug': 'discapacidad'})
DiscapacidadCreateView = make_create_view(Discapacidad,   DiscapacidadForm,   'discapacidad_list',     'Nueva Discapacidad',         'discapacidad_list')
DiscapacidadUpdateView = make_update_view(Discapacidad,   DiscapacidadForm,   'discapacidad_list',     'Editar',                    'discapacidad_list')
DiscapacidadDeleteView = make_delete_view(Discapacidad,   'discapacidad_list',     'Eliminar Discapacidad',      'discapacidad_list')

EnfermedadListView   = make_list_view(EnfermedadCronica,'catalogos/simple_list.html', 'Enfermedades Crónicas',  ['descripcion'],
                                      extra_ctx={'permite_eliminar': True, 'export_slug': 'enfermedad'})
EnfermedadCreateView = make_create_view(EnfermedadCronica,EnfermedadCronicaForm,'enfermedad_list',       'Nueva Enfermedad',           'enfermedad_list')
EnfermedadUpdateView = make_update_view(EnfermedadCronica,EnfermedadCronicaForm,'enfermedad_list',       'Editar',                    'enfermedad_list')
EnfermedadDeleteView = make_delete_view(EnfermedadCronica,'enfermedad_list',       'Eliminar Enfermedad',        'enfermedad_list')

PuebloListView   = make_list_view(Pueblo,        'catalogos/simple_list.html', 'Pueblos Indígenas',      ['descripcion'],
                                  extra_ctx={'permite_eliminar': True, 'export_slug': 'pueblo'})
PuebloCreateView = make_create_view(Pueblo,        PuebloForm,        'pueblo_list',           'Nuevo Pueblo',               'pueblo_list')
PuebloUpdateView = make_update_view(Pueblo,        PuebloForm,        'pueblo_list',           'Editar',                    'pueblo_list')
PuebloDeleteView = make_delete_view(Pueblo,        'pueblo_list',           'Eliminar Pueblo',            'pueblo_list')

MotivoBajaListView   = make_list_view(MotivoBaja,    'catalogos/simple_list.html', 'Motivos de Baja',        ['clave','descripcion'],
                                      extra_ctx={'export_slug': 'motivo_baja'})
MotivoBajaCreateView = make_create_view(MotivoBaja,    MotivoBajaForm,    'motivo_baja_list',      'Nuevo Motivo de Baja',       'motivo_baja_list')
MotivoBajaUpdateView = make_update_view(MotivoBaja,    MotivoBajaForm,    'motivo_baja_list',      'Editar',                    'motivo_baja_list')

IdiomaListView   = make_list_view(Idioma,        'catalogos/simple_list.html', 'Idiomas / Lenguas',      ['descripcion','familia_linguistica'],
                                  extra_ctx={'permite_eliminar': True, 'export_slug': 'idioma'})
IdiomaCreateView = make_create_view(Idioma,        IdiomaForm,        'idioma_list',           'Nuevo Idioma / Lengua',      'idioma_list')
IdiomaUpdateView = make_update_view(Idioma,        IdiomaForm,        'idioma_list',           'Editar',                    'idioma_list')
IdiomaDeleteView = make_delete_view(Idioma,        'idioma_list',           'Eliminar Idioma',            'idioma_list')

EstadoCivilListView   = make_list_view(EstadoCivil,   'catalogos/simple_list.html', 'Estados Civiles',        ['clave','descripcion'],
                                       extra_ctx={'permite_eliminar': True, 'export_slug': 'estado_civil'})
EstadoCivilCreateView = make_create_view(EstadoCivil,   EstadoCivilForm,   'estado_civil_list',     'Nuevo Estado Civil',         'estado_civil_list')
EstadoCivilUpdateView = make_update_view(EstadoCivil,   EstadoCivilForm,   'estado_civil_list',     'Editar',                    'estado_civil_list')
EstadoCivilDeleteView = make_delete_view(EstadoCivil,   'estado_civil_list',     'Eliminar Estado Civil',      'estado_civil_list')

PaisListView   = make_list_view(Pais,          'catalogos/simple_list.html', 'Países',                 ['nombre'],
                                extra_ctx={'permite_eliminar': True, 'export_slug': 'pais'})
PaisCreateView = make_create_view(Pais,          PaisForm,          'pais_list',             'Nuevo País',                'pais_list')
PaisUpdateView = make_update_view(Pais,          PaisForm,          'pais_list',             'Editar',                    'pais_list')
PaisDeleteView = make_delete_view(Pais,          'pais_list',             'Eliminar País',              'pais_list')

EntidadListView   = make_list_view(EntidadFederativa,'catalogos/simple_list.html', 'Entidades Federativas',  ['nombre','abreviatura'],
                                   extra_ctx={'permite_eliminar': True, 'export_slug': 'entidad'})
EntidadCreateView = make_create_view(EntidadFederativa,EntidadFederativaForm,'entidad_list',          'Nueva Entidad Federativa',   'entidad_list')
EntidadUpdateView = make_update_view(EntidadFederativa,EntidadFederativaForm,'entidad_list',          'Editar',                    'entidad_list')
EntidadDeleteView = make_delete_view(EntidadFederativa,'entidad_list',          'Eliminar Entidad Federativa', 'entidad_list')

class MunicipioListView(LoginRequiredMixin, ListView):
    model               = Municipio
    template_name       = 'catalogos/simple_list.html'
    context_object_name = 'registros'
    paginate_by         = 50

    def get_queryset(self):
        qs  = Municipio.objects.select_related('entidad')
        q   = self.request.GET.get('q', '')
        ent = self.request.GET.get('ent', '')
        if q:
            qs = qs.filter(Q(nombre__icontains=q) | Q(clave__icontains=q))
        if ent:
            qs = qs.filter(entidad_id=ent)
        return qs.order_by('entidad__nombre', 'nombre')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({'titulo': 'Municipios', 'q': self.request.GET.get('q', ''),
                    'entidades': EntidadFederativa.objects.all().order_by('nombre'),
                    'ent_sel': self.request.GET.get('ent', ''),
                    'permite_eliminar': True, 'export_slug': 'municipio'})
        return ctx

MunicipioCreateView = make_create_view(Municipio, MunicipioForm, 'municipio_list', 'Nuevo Municipio', 'municipio_list')
MunicipioUpdateView = make_update_view(Municipio, MunicipioForm, 'municipio_list', 'Editar Municipio', 'municipio_list')
MunicipioDeleteView = make_delete_view(Municipio, 'municipio_list', 'Eliminar Municipio', 'municipio_list')

SindicatoListView   = make_list_view(Sindicato,    'catalogos/simple_list.html', 'Sindicatos',             ['clave','descripcion'],
                                     extra_ctx={'permite_eliminar': True, 'export_slug': 'sindicato'})
SindicatoCreateView = make_create_view(Sindicato,    SindicatoForm,    'sindicato_list',        'Nuevo Sindicato',            'sindicato_list')
SindicatoUpdateView = make_update_view(Sindicato,    SindicatoForm,    'sindicato_list',        'Editar',                    'sindicato_list')
SindicatoDeleteView = make_delete_view(Sindicato,    'sindicato_list',        'Eliminar Sindicato',         'sindicato_list')

class InmuebleListView(LoginRequiredMixin, ListView):
    model               = Inmueble
    template_name       = 'catalogos/simple_list.html'
    context_object_name = 'registros'
    paginate_by         = 30

    def get_queryset(self):
        qs  = filtrar_por_dependencia(Inmueble.objects.select_related('dependencia'), self.request.user)
        q   = self.request.GET.get('q', '')
        if q:
            qs = qs.filter(Q(clave__icontains=q) | Q(descripcion__icontains=q))
        return qs.order_by('dependencia__clave', 'clave')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({'titulo': 'Inmuebles', 'q': self.request.GET.get('q', ''), 'export_slug': 'inmueble'})
        return ctx

InmuebleCreateView = make_create_view(Inmueble, InmuebleForm, 'inmueble_list', 'Nuevo Inmueble', 'inmueble_list',
                                      mixins=(DependenciaFormRestrictMixin,))
InmuebleUpdateView = make_update_view(Inmueble, InmuebleForm, 'inmueble_list', 'Editar Inmueble', 'inmueble_list',
                                      mixins=(DependenciaFormRestrictMixin, DependenciaScopedMixin))


# ── Descarga de plantillas ────────────────────────────────────────────────────
@login_required
def descargar_catalogo(request, catalogo):
    nombres = {
        'fuente':      'Catalogo_Fuente.xlsx',
        'dependencia': 'Catalogo_Dependencia.xlsx',
        'categoria':   'Catalogo_Categoria.xlsx',
        'unidades':    'Catalogo_Unidades_Admvas.xlsx',
        'programas':   'Catalogo_Programas.xlsx',
        'proyectos':   'Catalogo_Proyectos.xlsx',
        'puestos':     'Catalogo_Plazas.xlsx',
    }
    if catalogo not in nombres:
        raise Http404("Plantilla no encontrada.")
    nombre_archivo = nombres[catalogo]
    base  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ruta  = os.path.join(base, nombre_archivo)
    if not os.path.exists(ruta):
        raise Http404(f"El archivo {nombre_archivo} no está disponible.")
    response = FileResponse(
        open(ruta, 'rb'),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


# ── Exportar TODOS los registros de un catálogo a Excel ───────────────────────
@login_required
def exportar_catalogo_excel(request, catalogo):
    """A diferencia de 'descargar_catalogo' (plantilla en blanco para
    importar), esto exporta los registros que YA existen en la tabla —
    completos, sin el límite de 10 de la paginación de la lista. Si el
    catálogo tiene dueño por dependencia (Unidades, Programas, Proyectos,
    Inmuebles...), un usuario no-administrador solo exporta los de la suya,
    igual que en el resto del sistema."""
    from .exportador import CATALOGOS_EXPORT
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    cfg = CATALOGOS_EXPORT.get(catalogo)
    if not cfg:
        raise Http404("Catálogo no disponible para exportar.")

    qs = cfg['model'].objects.all()
    if cfg.get('select_related'):
        qs = qs.select_related(*cfg['select_related'])
    if cfg.get('dependencia_lookup'):
        qs = filtrar_por_dependencia(qs, request.user, cfg['dependencia_lookup'])
    qs = qs.order_by(*cfg['order_by'])

    columnas = cfg['columnas']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cfg['titulo'][:31]  # límite de Excel para nombres de hoja

    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col, (encabezado, _f) in enumerate(columnas, 1):
        celda = ws.cell(row=1, column=col, value=encabezado)
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal='center')

    for fila_num, obj in enumerate(qs, 2):
        for col, (_h, extraer) in enumerate(columnas, 1):
            ws.cell(row=fila_num, column=col, value=extraer(obj))

    for i in range(1, len(columnas) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="RESP_{cfg["titulo"].replace(" ", "_")}.xlsx"'
    wb.save(response)
    return response


# ── Importación de catálogos desde Excel ──────────────────────────────────────
def _importar_catalogo_en_segundo_plano(importacion_pk, catalogo, ruta_tmp):
    """Corre el importador del catálogo fuera del ciclo request/response, en
    un hilo aparte: Plazas puede traer decenas de miles de filas (con varias
    consultas de catálogo por fila), tiempo de sobra para chocar con el
    504 Gateway Timeout de nginx si se corriera dentro del request. La
    vista que lanza esto ya dejó la ImportacionCatalogo en 'procesando'
    antes de arrancar el hilo, así que la pantalla de detalle puede
    auto-refrescarse hasta que este hilo la actualice con el resultado."""
    from django.db import connection
    from .importador import IMPORTADORES

    funcion = IMPORTADORES[catalogo][0]
    try:
        importacion = ImportacionCatalogo.objects.get(pk=importacion_pk)
        try:
            ruta_abs = default_storage.path(ruta_tmp)
            resultado = funcion(ruta_abs)
            importacion.creados = resultado['creados']
            importacion.actualizados = resultado['actualizados']
            importacion.errores = resultado['errores']
            importacion.total = resultado['total']
            importacion.log = resultado['log']
            importacion.estado = 'completado'
            importacion.save()
        except Exception as e:
            importacion.errores = 1
            importacion.log = [f'No se pudo procesar el archivo: {e}']
            importacion.estado = 'error'
            importacion.save()
        finally:
            default_storage.delete(ruta_tmp)
    finally:
        connection.close()


@login_required
@admin_requerido
def importar_catalogo(request, catalogo):
    """La carga masiva de catálogos abarca claves de todo el sistema (no de
    una sola dependencia), así que queda reservada a administradores."""
    from .importador import IMPORTADORES

    if catalogo not in IMPORTADORES:
        raise Http404("Catálogo no disponible para importar.")
    funcion, url_lista, titulo = IMPORTADORES[catalogo]

    # Si ya hay una importación de este catálogo corriendo, no tiene sentido
    # mostrar el formulario (ni menos lanzar una segunda en paralelo contra
    # las mismas tablas) — se retoma directo su pantalla de estado.
    en_curso = ImportacionCatalogo.objects.filter(catalogo=catalogo, estado='procesando').order_by('-fecha').first()
    if en_curso:
        if request.method == 'POST':
            messages.warning(
                request,
                'Ya hay una importación de este catálogo en curso — espere a que termine antes de subir otra.'
            )
        return redirect('importacion_detalle', pk=en_curso.pk)

    if request.method == 'POST':
        form = ImportarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            ruta_tmp = default_storage.save(f'tmp_importaciones/{archivo.name}', archivo)
            importacion = ImportacionCatalogo.objects.create(
                catalogo=catalogo, nombre_archivo=archivo.name, usuario=request.user,
            )
            hilo = threading.Thread(
                target=_importar_catalogo_en_segundo_plano,
                args=(importacion.pk, catalogo, ruta_tmp),
                daemon=True,
            )
            hilo.start()
            messages.info(
                request,
                'Archivo recibido, importándose en segundo plano — esta página se '
                'actualizará sola cuando termine.'
            )
            return redirect('importacion_detalle', pk=importacion.pk)
    else:
        form = ImportarExcelForm()

    historial = ImportacionCatalogo.objects.filter(catalogo=catalogo).select_related('usuario').order_by('-fecha')[:15]

    return render(request, 'catalogos/importar_form.html', {
        'form': form,
        'titulo': f'Importar {titulo}',
        'catalogo': catalogo,
        'url_lista': url_lista,
        'historial': historial,
    })


@login_required
@admin_requerido
def importacion_detalle(request, pk):
    from .importador import IMPORTADORES
    importacion = get_object_or_404(ImportacionCatalogo, pk=pk)
    _, url_lista, titulo = IMPORTADORES.get(importacion.catalogo, (None, 'catalogo_index', importacion.catalogo))
    return render(request, 'catalogos/importar_detalle.html', {
        'importacion': importacion,
        'titulo': f'Importar {titulo}',
        'url_lista': url_lista,
    })
