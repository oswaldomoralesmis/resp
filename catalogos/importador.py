# -*- coding: utf-8 -*-
"""Importadores de catálogos desde Excel (plantillas Catalogo_*.xlsx).

Cada función toma la ruta de un .xlsx subido por el usuario, procesa las
filas desde la 4 (row 1-3 son título/leyenda/encabezados) y devuelve:
    {'creados': int, 'actualizados': int, 'errores': int, 'total': int, 'log': [str, ...]}
"""
from openpyxl import load_workbook

from .models import (
    Dependencia, UnidadAdministrativa, Programa, Proyecto, FuenteFinanciamiento,
    Categoria, TipoContratacion, NivelEstructura, EstatusPlaza, CentroTrabajo,
)
from servidores.models import Puesto


def sd(v):
    """String seguro: recorta espacios, colapsa None a ''."""
    if v is None:
        return ''
    return str(v).strip()


def _es_fila_ejemplo(valores):
    """La plantilla trae una fila de ejemplo (fila 4) con celdas 'Ej: ...'."""
    return any(sd(v).upper().startswith('EJ:') for v in valores if v is not None)


def _cargar_hoja(ruta):
    wb = load_workbook(ruta, data_only=True)
    return wb[wb.sheetnames[0]]


def importar_fuentes(ruta):
    ws = _cargar_hoja(ruta)
    creados = actualizados = errores = total = 0
    log = []
    for num_fila, fila in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not any(c is not None for c in fila):
            continue
        if _es_fila_ejemplo(fila):
            continue
        total += 1
        clave = sd(fila[0]).upper()
        desc = sd(fila[1]) if len(fila) > 1 else ''
        if not clave or not desc:
            errores += 1
            log.append(f'Fila {num_fila}: OMITIDA — clave o descripción vacía')
            continue
        obj, creado = FuenteFinanciamiento.objects.update_or_create(
            clave=clave, defaults={'descripcion': desc}
        )
        if creado:
            creados += 1
        else:
            actualizados += 1
    log.insert(0, f'{creados} creadas, {actualizados} actualizadas, {errores} con error, {total} filas procesadas.')
    return {'creados': creados, 'actualizados': actualizados, 'errores': errores, 'total': total, 'log': log}


def importar_dependencias(ruta):
    ws = _cargar_hoja(ruta)
    creados = actualizados = errores = total = 0
    log = []
    for num_fila, fila in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not any(c is not None for c in fila):
            continue
        if _es_fila_ejemplo(fila):
            continue
        total += 1
        clave = sd(fila[0]).upper()
        desc = sd(fila[1]) if len(fila) > 1 else ''
        if not clave or not desc:
            errores += 1
            log.append(f'Fila {num_fila}: OMITIDA — clave o descripción vacía')
            continue
        obj, creado = Dependencia.objects.update_or_create(
            clave=clave, defaults={'descripcion': desc}
        )
        if creado:
            creados += 1
        else:
            actualizados += 1
    log.insert(0, f'{creados} creadas, {actualizados} actualizadas, {errores} con error, {total} filas procesadas.')
    return {'creados': creados, 'actualizados': actualizados, 'errores': errores, 'total': total, 'log': log}


def importar_unidades(ruta):
    ws = _cargar_hoja(ruta)
    creados = actualizados = errores = total = 0
    log = []
    for num_fila, fila in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not any(c is not None for c in fila):
            continue
        if _es_fila_ejemplo(fila):
            continue
        total += 1
        dep_clave = sd(fila[0]).upper()
        clave = sd(fila[1]).upper() if len(fila) > 1 else ''
        desc = sd(fila[2]) if len(fila) > 2 else ''
        if not dep_clave or not clave or not desc:
            errores += 1
            log.append(f'Fila {num_fila}: OMITIDA — dependencia, clave o descripción vacía')
            continue
        dep = Dependencia.objects.filter(clave=dep_clave).first()
        if not dep:
            errores += 1
            log.append(f'Fila {num_fila}: ERROR — dependencia "{dep_clave}" no encontrada')
            continue
        obj, creado = UnidadAdministrativa.objects.update_or_create(
            dependencia=dep, clave=clave, defaults={'descripcion': desc}
        )
        if creado:
            creados += 1
        else:
            actualizados += 1
    log.insert(0, f'{creados} creadas, {actualizados} actualizadas, {errores} con error, {total} filas procesadas.')
    return {'creados': creados, 'actualizados': actualizados, 'errores': errores, 'total': total, 'log': log}


def importar_programas(ruta):
    ws = _cargar_hoja(ruta)
    creados = actualizados = errores = total = 0
    log = []
    for num_fila, fila in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not any(c is not None for c in fila):
            continue
        if _es_fila_ejemplo(fila):
            continue
        total += 1
        dep_clave = sd(fila[0]).upper()
        uni_clave = sd(fila[1]).upper() if len(fila) > 1 else ''
        clave = sd(fila[2]).upper() if len(fila) > 2 else ''
        desc = sd(fila[3]) if len(fila) > 3 else ''
        if not dep_clave or not uni_clave or not clave or not desc:
            errores += 1
            log.append(f'Fila {num_fila}: OMITIDA — faltan datos obligatorios')
            continue
        dep = Dependencia.objects.filter(clave=dep_clave).first()
        uni = UnidadAdministrativa.objects.filter(dependencia=dep, clave=uni_clave).first() if dep else None
        if not dep or not uni:
            errores += 1
            log.append(f'Fila {num_fila}: ERROR — dependencia "{dep_clave}" o unidad "{uni_clave}" no encontrada')
            continue
        obj, creado = Programa.objects.update_or_create(
            unidad=uni, clave=clave, defaults={'dependencia': dep, 'descripcion': desc}
        )
        if creado:
            creados += 1
        else:
            actualizados += 1
    log.insert(0, f'{creados} creados, {actualizados} actualizados, {errores} con error, {total} filas procesadas.')
    return {'creados': creados, 'actualizados': actualizados, 'errores': errores, 'total': total, 'log': log}


def importar_proyectos(ruta):
    """La misma CLAVE_PROYECTO puede repetirse en varias filas (una por cada
    programa/unidad); todas se fusionan en un solo Proyecto vía Proyecto.programas."""
    ws = _cargar_hoja(ruta)
    creados = actualizados = errores = total = 0
    log = []
    for num_fila, fila in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not any(c is not None for c in fila):
            continue
        if _es_fila_ejemplo(fila):
            continue
        # La fila de nota final ("NOTA: ...") no trae dependencia/clave, se ignora.
        if not (len(fila) > 4 and fila[1] and fila[4]):
            continue
        total += 1
        dep_clave = sd(fila[1]).upper()
        uni_clave = sd(fila[2]).upper() if len(fila) > 2 else ''
        prg_clave = sd(fila[3]).upper() if len(fila) > 3 else ''
        clave = sd(fila[4]).upper()
        desc = sd(fila[5]) if len(fila) > 5 else ''
        if not dep_clave or not uni_clave or not prg_clave or not clave or not desc:
            errores += 1
            log.append(f'Fila {num_fila}: OMITIDA — faltan datos obligatorios')
            continue
        dep = Dependencia.objects.filter(clave=dep_clave).first()
        uni = UnidadAdministrativa.objects.filter(dependencia=dep, clave=uni_clave).first() if dep else None
        prg = Programa.objects.filter(unidad=uni, clave=prg_clave).first() if uni else None
        if not dep or not uni or not prg:
            errores += 1
            log.append(
                f'Fila {num_fila}: ERROR — dependencia "{dep_clave}", unidad "{uni_clave}" '
                f'o programa "{prg_clave}" no encontrada'
            )
            continue
        obj, creado = Proyecto.objects.update_or_create(
            dependencia=dep, clave=clave, defaults={'descripcion': desc}
        )
        obj.programas.add(prg)
        if creado:
            creados += 1
        else:
            actualizados += 1
    log.insert(0, f'{creados} creados, {actualizados} actualizados, {errores} con error, {total} filas procesadas.')
    return {'creados': creados, 'actualizados': actualizados, 'errores': errores, 'total': total, 'log': log}


def importar_puestos(ruta):
    """Alta/actualización masiva de plazas (solo datos de la plaza: ubicación
    presupuestal, categoría, contratación, estatus, CCT y plaza jefe). No toca
    servidor_actual: eso lo asigna la carga quincenal de Información Básica."""
    ws = _cargar_hoja(ruta)
    creados = actualizados = errores = total = 0
    log = []
    for num_fila, fila in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not any(c is not None for c in fila):
            continue
        if _es_fila_ejemplo(fila):
            continue
        total += 1
        dep_clave   = sd(fila[0]).upper()
        uni_clave   = sd(fila[1]).upper() if len(fila) > 1 else ''
        prg_clave   = sd(fila[2]).upper() if len(fila) > 2 else ''
        pry_clave   = sd(fila[3]).upper() if len(fila) > 3 else ''
        id_plaza    = sd(fila[4]) if len(fila) > 4 else ''
        cat_clave   = sd(fila[5]).upper() if len(fila) > 5 else ''
        tc_clave    = sd(fila[6]).upper() if len(fila) > 6 else ''
        ne_clave    = sd(fila[7]).upper() if len(fila) > 7 else ''
        ep_clave    = sd(fila[8]).upper() if len(fila) > 8 else ''
        cct_clave   = sd(fila[9]).upper() if len(fila) > 9 else ''
        id_plaza_jefe = sd(fila[10]) if len(fila) > 10 else ''

        if not (dep_clave and uni_clave and prg_clave and pry_clave and id_plaza and cat_clave and tc_clave):
            errores += 1
            log.append(f'Fila {num_fila}: OMITIDA — faltan datos obligatorios')
            continue

        dep = Dependencia.objects.filter(clave=dep_clave).first()
        uni = UnidadAdministrativa.objects.filter(dependencia=dep, clave=uni_clave).first() if dep else None
        prg = Programa.objects.filter(unidad=uni, clave=prg_clave).first() if uni else None
        pry = Proyecto.objects.filter(dependencia=dep, clave=pry_clave).first() if dep else None
        cat = Categoria.objects.filter(clave=cat_clave).first()
        tc  = TipoContratacion.objects.filter(clave=tc_clave).first()
        if not (dep and uni and prg and pry and cat and tc):
            errores += 1
            log.append(
                f'Fila {num_fila}: ERROR — dependencia "{dep_clave}", unidad "{uni_clave}", '
                f'programa "{prg_clave}", proyecto "{pry_clave}", categoría "{cat_clave}" o '
                f'tipo de contratación "{tc_clave}" no encontrados'
            )
            continue
        ne  = NivelEstructura.objects.filter(clave=ne_clave).first() if ne_clave else None
        ep  = EstatusPlaza.objects.filter(clave=ep_clave).first() if ep_clave else None
        cct = CentroTrabajo.objects.filter(clave=cct_clave).first() if cct_clave else None

        obj, creado = Puesto.objects.update_or_create(
            id_plaza=id_plaza,
            defaults={
                'proyecto': pry, 'programa': prg, 'unidad': uni, 'categoria': cat,
                'nombramiento': tc, 'nivel_estructura': ne, 'estatus_plaza': ep,
                'cct': cct, 'id_plaza_jefe': id_plaza_jefe,
            }
        )
        if creado:
            creados += 1
        else:
            actualizados += 1
    log.insert(0, f'{creados} creados, {actualizados} actualizados, {errores} con error, {total} filas procesadas.')
    return {'creados': creados, 'actualizados': actualizados, 'errores': errores, 'total': total, 'log': log}


IMPORTADORES = {
    'fuente':      (importar_fuentes,      'fuente_list',      'Fuentes de Financiamiento'),
    'dependencia': (importar_dependencias, 'dependencia_list', 'Dependencias'),
    'unidades':    (importar_unidades,     'unidad_list',      'Unidades Administrativas'),
    'programas':   (importar_programas,    'programa_list',    'Programas'),
    'proyectos':   (importar_proyectos,    'proyecto_list',    'Proyectos'),
    'puestos':     (importar_puestos,      'puesto_list',      'Plazas'),
}
