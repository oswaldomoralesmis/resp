# -*- coding: utf-8 -*-
from datetime import date
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Count, ExpressionWrapper, F, IntegerField, Q, Value
from django.db.models.functions import Cast, Substr
from django.http import HttpResponse
from servidores.models import ServidorPublico, InformacionBasica, BajaServidorPublico, Puesto, DiscapacidadServidor
from catalogos.models import Dependencia
from usuarios.mixins import filtrar_por_dependencia, permiso_requerido
from .pdf import generar_pdf_paridad, generar_pdf_ocupacion, generar_pdf_discapacidad
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


@login_required
def reporte_index(request):
    return render(request, 'reportes/index.html', {'titulo': 'Reportes y Estadísticas'})


@login_required
@permiso_requerido('reporte_padron')
def reporte_padron(request):
    qs = filtrar_por_dependencia(
        ServidorPublico.objects.filter(activo=True).select_related('entidad_nacimiento', 'pais_nacimiento', 'sindicato'),
        request.user, 'informacion_basica__dependencia_id',
    ).distinct()
    dep = request.GET.get('dependencia', '')
    if dep:
        qs = qs.filter(informacion_basica__dependencia_id=dep, informacion_basica__activo=True)
    context = {
        'servidores': qs[:500],
        'total': qs.count(),
        'dependencias': filtrar_por_dependencia(Dependencia.objects.all(), request.user, 'pk'),
        'titulo': 'Padrón de Servidores Públicos',
    }
    return render(request, 'reportes/padron.html', context)


@login_required
@permiso_requerido('reporte_bajas')
def reporte_bajas(request):
    bajas = filtrar_por_dependencia(
        BajaServidorPublico.objects.select_related('servidor', 'dependencia', 'motivo_baja'), request.user
    )
    carga_pk = request.GET.get('carga', '')
    if carga_pk:
        bajas = bajas.filter(carga_origen_id=carga_pk)
    q = request.GET.get('q', '')
    if q:
        bajas = bajas.filter(
            Q(servidor__rfc__icontains=q) | Q(servidor__nombre__icontains=q) |
            Q(servidor__primer_apellido__icontains=q) | Q(servidor__expediente__icontains=q)
        )
    bajas = bajas.order_by('-fecha_baja')
    return render(request, 'reportes/bajas.html', {
        'bajas': bajas, 'titulo': 'Bajas de Servidores Públicos',
        'q': q, 'carga_filtro': carga_pk,
    })


@login_required
@permiso_requerido('reporte_declaracion')
def reporte_declaracion(request):
    registros = filtrar_por_dependencia(
        InformacionBasica.objects.filter(activo=True, oblig_declaracion='S')
        .select_related('servidor', 'dependencia', 'tipo_declaracion'),
        request.user,
    ).order_by('dependencia', 'servidor')
    return render(request, 'reportes/declaracion.html', {
        'registros': registros, 'titulo': 'Declaración Patrimonial - Sujetos Obligados'
    })


@login_required
@permiso_requerido('reporte_entrega_recepcion')
def reporte_entrega_recepcion(request):
    registros = filtrar_por_dependencia(
        InformacionBasica.objects.filter(activo=True, oblig_entrega_recepcion='S').select_related('servidor', 'dependencia'),
        request.user,
    ).order_by('dependencia', 'servidor')
    return render(request, 'reportes/entrega_recepcion.html', {
        'registros': registros, 'titulo': 'Padrón - Acta de Entrega-Recepción'
    })


@login_required
@permiso_requerido('reporte_compatibilidad')
def reporte_compatibilidad(request):
    servidores_doble = filtrar_por_dependencia(
        ServidorPublico.objects.filter(activo=True, tiene_otra_plaza='S'),
        request.user, 'informacion_basica__dependencia_id',
    ).distinct()
    return render(request, 'reportes/compatibilidad.html', {
        'servidores': servidores_doble, 'titulo': 'Compatibilidad de Horarios'
    })


SEXO_LABELS = {'MASCULINO': 'Masculino', 'FEMENINO': 'Femenino', 'OTRO': 'Otro'}


def _distribucion(qs, campo, top=None):
    """Cuenta registros de qs agrupados por 'campo' (ruta de FK, p.ej.
    'programa__descripcion'), como lista de {'label': ..., 'total': ...}
    ordenada de mayor a menor. 'top' recorta a los N más frecuentes."""
    filas = qs.values(campo).annotate(total=Count('id')).order_by('-total')
    if top:
        filas = filas[:top]
    return [{'label': f[campo] or 'Sin dato', 'total': f['total']} for f in filas]


def _distribucion_sexo_con_pct(qs, campo_sexo='sexo'):
    """Como _distribucion pero solo para el campo de sexo, agregando el
    porcentaje de cada valor sobre el total — para el resumen tipo
    'pastel 52%/48%' de los monitores de paridad/discapacidad."""
    filas = list(qs.values(campo_sexo).annotate(total=Count('id')).order_by('-total'))
    total = sum(f['total'] for f in filas)
    return [
        {
            'label': SEXO_LABELS.get(f[campo_sexo], f[campo_sexo] or 'Sin dato'),
            'total': f['total'],
            'porcentaje': round(f['total'] * 100 / total, 1) if total else 0,
        }
        for f in filas
    ]


def _cruce_por_sexo(qs, campo, campo_sexo='sexo', top=None):
    """Tabla cruzada 'campo' (p.ej. nombramiento__descripcion) × sexo: para
    cada valor de 'campo' cuenta hombres/mujeres/otro y el % de mujeres —
    el mismo cruce que usan los monitores de paridad/discapacidad del RUSP
    federal (tipo de personal según sexo, tipo de contratación según
    sexo, etc.)."""
    filas = qs.values(campo, campo_sexo).annotate(total=Count('id'))
    acumulado = {}
    for f in filas:
        label = f[campo] or 'Sin dato'
        c = acumulado.setdefault(label, {'hombres': 0, 'mujeres': 0, 'otro': 0})
        if f[campo_sexo] == 'MASCULINO':
            c['hombres'] += f['total']
        elif f[campo_sexo] == 'FEMENINO':
            c['mujeres'] += f['total']
        else:
            c['otro'] += f['total']
    resultado = []
    for label, c in acumulado.items():
        total = c['hombres'] + c['mujeres'] + c['otro']
        resultado.append({
            'label': label, 'hombres': c['hombres'], 'mujeres': c['mujeres'], 'otro': c['otro'],
            'total': total,
            'pct_hombres': round(c['hombres'] * 100 / total, 1) if total else 0,
            'pct_mujeres': round(c['mujeres'] * 100 / total, 1) if total else 0,
        })
    resultado.sort(key=lambda x: -x['total'])
    return resultado[:top] if top else resultado


@login_required
@permiso_requerido('reporte_estadisticas')
def reporte_estadisticas(request):
    info_visible = filtrar_por_dependencia(InformacionBasica.objects.filter(activo=True), request.user)
    puestos_visibles = filtrar_por_dependencia(Puesto.objects.all(), request.user, 'proyecto__dependencia')
    servidores_visibles = filtrar_por_dependencia(
        ServidorPublico.objects.filter(activo=True), request.user, 'informacion_basica__dependencia_id'
    ).distinct()
    bajas_visibles = filtrar_por_dependencia(BajaServidorPublico.objects.all(), request.user)

    ocupadas = list(info_visible.values('estatus_plaza__descripcion').annotate(total=Count('id')))
    vacantes = list(
        puestos_visibles.filter(servidor_actual__isnull=True).values('estatus_plaza__descripcion').annotate(total=Count('id'))
    )

    por_estatus_map = {}
    for item in ocupadas + vacantes:
        label = item['estatus_plaza__descripcion'] or 'Sin estatus'
        por_estatus_map[label] = por_estatus_map.get(label, 0) + item['total']

    total_info = info_visible.count()

    def porcentaje_obligados(campo):
        obligados = info_visible.filter(**{campo: 'S'}).count()
        pct = round(obligados * 100 / total_info, 1) if total_info else 0
        return {'total': obligados, 'porcentaje': pct}

    por_sexo_raw = servidores_visibles.values('sexo').annotate(total=Count('id')).order_by('-total')

    stats = {
        'total_activos': servidores_visibles.count(),
        'total_bajas': bajas_visibles.count(),
        'total_info': total_info,

        # Institucional
        'por_dependencia': _distribucion(info_visible, 'dependencia__descripcion', top=15),
        'por_programa': _distribucion(info_visible, 'programa__descripcion', top=15),
        'por_fuente_financiamiento': _distribucion(info_visible, 'fuente_financiamiento__descripcion'),
        'por_area': _distribucion(info_visible, 'area__descripcion', top=15),

        # Puesto / función
        'por_estatus': [
            {'label': label, 'total': total}
            for label, total in sorted(por_estatus_map.items(), key=lambda item: (-item[1], item[0]))
        ],
        'por_contratacion': _distribucion(info_visible, 'nombramiento__descripcion'),
        'por_categoria': _distribucion(info_visible, 'categoria__descripcion', top=15),
        'por_nivel_estructura': _distribucion(info_visible, 'nivel_estructura__descripcion'),
        'por_tipo_funcion': _distribucion(info_visible, 'tipo_funcion__descripcion'),
        'por_tipo_personal': _distribucion(info_visible, 'tipo_personal__descripcion'),
        'por_cct': _distribucion(info_visible, 'cct__nombre', top=15),

        # Personal
        'por_sexo': [
            {'label': SEXO_LABELS.get(f['sexo'], f['sexo'] or 'Sin dato'), 'total': f['total']}
            for f in por_sexo_raw
        ],
        'por_estado_civil': _distribucion(servidores_visibles, 'estado_civil__descripcion'),
        'por_entidad_nacimiento': _distribucion(servidores_visibles, 'entidad_nacimiento__nombre', top=15),
        'por_sindicato': _distribucion(servidores_visibles, 'sindicato__descripcion'),

        # Cumplimiento (sobre InformacionBasica activa)
        'obligados_declaracion': porcentaje_obligados('oblig_declaracion'),
        'obligados_entrega_recepcion': porcentaje_obligados('oblig_entrega_recepcion'),
        'obligados_rendir_cuentas': porcentaje_obligados('oblig_rendir_cuentas'),

        # Bajas
        'por_motivo_baja': _distribucion(bajas_visibles, 'motivo_baja__descripcion'),
    }
    return render(request, 'reportes/estadisticas.html', {'stats': stats, 'titulo': 'Estadísticas Generales'})


MESES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
    7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
}


def _evolucion_por_sexo(request, por='mes'):
    """Evolución histórica del % de hombres/mujeres por mes o por año,
    usando TODO el histórico de InformacionBasica (vigente e inactiva) —
    cada registro ya es una fotografía fija de cómo estaba la persona en
    esa quincena (ver 'fotografía por periodo'), así que sirve tal cual
    para reconstruir el pasado sin que una corrección posterior en
    Padrón lo altere. 'quincena' se guarda como 'AAAA-QQ' (QQ 01-24, dos
    quincenas por mes en orden: 01-02 enero, 03-04 febrero, ...); dentro
    de cada mes/año se toma solo la quincena MÁS RECIENTE de cada
    servidor, para no contarlo dos veces solo porque tuvo dos quincenas
    en el mismo periodo."""
    info = filtrar_por_dependencia(InformacionBasica.objects.all(), request.user)
    info = info.exclude(quincena='').filter(quincena__regex=r'^\d{4}-\d{2}$').annotate(
        anio=Cast(Substr('quincena', 1, 4), IntegerField()),
        quincena_num=Cast(Substr('quincena', 6, 2), IntegerField()),
    )

    if por == 'mes':
        info = info.annotate(
            periodo_num=ExpressionWrapper((F('quincena_num') + Value(1)) / Value(2), output_field=IntegerField())
        )
        campos_distinct = ['servidor_id', 'anio', 'periodo_num']
    else:
        campos_distinct = ['servidor_id', 'anio']

    # '-id' al final es puro desempate determinista: si el mismo servidor
    # trae más de una fila para la MISMA quincena (p.ej. dos plazas por
    # compatibilidad, cada una con su propia fotografía) y esas filas no
    # coinciden exactamente en sexo, sin un desempate fijo Postgres podía
    # devolver una u otra fila de forma arbitraria en cada ejecución,
    # haciendo que el reporte diera un número distinto cada vez que se
    # abriera. Con esto el resultado es siempre el mismo (no resuelve la
    # inconsistencia de origen entre las dos filas, pero sí hace que el
    # reporte sea reproducible).
    ultimo = info.order_by(*campos_distinct, '-quincena_num', '-id').distinct(*campos_distinct)

    acumulado = {}
    for row in ultimo:
        clave = (row.anio, row.periodo_num) if por == 'mes' else (row.anio,)
        c = acumulado.setdefault(clave, {'hombres': 0, 'mujeres': 0, 'otro': 0})
        if row.sexo == 'MASCULINO':
            c['hombres'] += 1
        elif row.sexo == 'FEMENINO':
            c['mujeres'] += 1
        else:
            c['otro'] += 1

    resultado = []
    for clave in sorted(acumulado.keys()):
        c = acumulado[clave]
        total = c['hombres'] + c['mujeres'] + c['otro']
        label = f'{MESES.get(clave[1], clave[1])} {clave[0]}' if por == 'mes' else str(clave[0])
        resultado.append({
            'label': label, 'hombres': c['hombres'], 'mujeres': c['mujeres'], 'otro': c['otro'], 'total': total,
            'pct_hombres': round(c['hombres'] * 100 / total, 1) if total else 0,
            'pct_mujeres': round(c['mujeres'] * 100 / total, 1) if total else 0,
        })
    return resultado


def _stats_paridad(request, top_categoria=15, top_dependencia=15):
    """Arma los datos del Monitor de Paridad. 'top_categoria'/'top_dependencia'
    en None trae TODAS las filas (para el PDF); un número recorta a las N
    más numerosas (para no saturar la pantalla)."""
    servidores_visibles = filtrar_por_dependencia(
        ServidorPublico.objects.filter(activo=True), request.user, 'informacion_basica__dependencia_id'
    ).distinct()
    info_visible = filtrar_por_dependencia(InformacionBasica.objects.filter(activo=True), request.user)

    return {
        'total_personas': servidores_visibles.count(),
        'por_sexo_personas': _distribucion_sexo_con_pct(servidores_visibles),
        'por_contratacion': _cruce_por_sexo(info_visible, 'nombramiento__descripcion'),
        'por_categoria': _cruce_por_sexo(info_visible, 'categoria__descripcion', top=top_categoria),
        'por_dependencia': _cruce_por_sexo(info_visible, 'dependencia__descripcion', top=top_dependencia),
        'evolucion_mensual': _evolucion_por_sexo(request, por='mes'),
        'evolucion_anual': _evolucion_por_sexo(request, por='anio'),
    }


@login_required
@permiso_requerido('reporte_paridad')
def reporte_paridad(request):
    """Monitor de Paridad de Género, al estilo del RUSP federal: total de
    personas por sexo, y esa misma distribución cruzada por tipo de
    contratación, categoría y dependencia. Los cruces cuentan PLAZAS
    ocupadas (InformacionBasica activa) — quien tenga más de una plaza
    puede contar más de una vez ahí, igual que en el reporte federal
    ('puestos ocupados por tipo de personal'); el total de personas de
    arriba sí es de personas físicas distintas."""
    stats = _stats_paridad(request)
    return render(request, 'reportes/paridad.html', {'stats': stats, 'titulo': 'Monitor de Paridad de Género'})


@login_required
@permiso_requerido('reporte_paridad')
def reporte_paridad_pdf(request):
    stats = _stats_paridad(request, top_categoria=None, top_dependencia=None)
    pdf = generar_pdf_paridad(stats)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Monitor_Paridad_RESP.pdf"'
    return response


def _cruce_ocupacion(puestos_visibles, campo, top=None):
    filas = puestos_visibles.values(campo).annotate(
        ocupadas=Count('id', filter=Q(servidor_actual__isnull=False)),
        vacantes=Count('id', filter=Q(servidor_actual__isnull=True)),
    )
    resultado = []
    for f in filas:
        label = f[campo] or 'Sin dato'
        total = f['ocupadas'] + f['vacantes']
        resultado.append({
            'label': label, 'ocupadas': f['ocupadas'], 'vacantes': f['vacantes'], 'total': total,
            'pct_ocupadas': round(f['ocupadas'] * 100 / total, 1) if total else 0,
        })
    resultado.sort(key=lambda x: -x['total'])
    return resultado[:top] if top else resultado


def _stats_ocupacion(request, top_categoria=15, top_dependencia=15):
    puestos_visibles = filtrar_por_dependencia(Puesto.objects.all(), request.user, 'proyecto__dependencia')
    total = puestos_visibles.count()
    ocupadas = puestos_visibles.filter(servidor_actual__isnull=False).count()
    vacantes = total - ocupadas
    return {
        'total': total,
        'ocupadas': ocupadas,
        'vacantes': vacantes,
        'pct_ocupadas': round(ocupadas * 100 / total, 1) if total else 0,
        'pct_vacantes': round(vacantes * 100 / total, 1) if total else 0,
        'por_contratacion': _cruce_ocupacion(puestos_visibles, 'nombramiento__descripcion'),
        'por_categoria': _cruce_ocupacion(puestos_visibles, 'categoria__descripcion', top=top_categoria),
        'por_dependencia': _cruce_ocupacion(puestos_visibles, 'proyecto__dependencia__descripcion', top=top_dependencia),
    }


@login_required
@permiso_requerido('reporte_ocupacion')
def reporte_ocupacion(request):
    """Plazas ocupadas vs. vacantes con porcentaje, y ese mismo cruce por
    tipo de contratación, categoría y dependencia — como el 'Monitor de
    Estadística de Puestos' del RUSP federal."""
    stats = _stats_ocupacion(request)
    return render(request, 'reportes/ocupacion.html', {'stats': stats, 'titulo': 'Ocupación de Plazas'})


@login_required
@permiso_requerido('reporte_ocupacion')
def reporte_ocupacion_pdf(request):
    stats = _stats_ocupacion(request, top_categoria=None, top_dependencia=None)
    pdf = generar_pdf_ocupacion(stats)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Ocupacion_Plazas_RESP.pdf"'
    return response


RANGOS_EDAD = [
    (24, 'Hasta 24'), (30, '25 a 30'), (35, '31 a 35'), (40, '36 a 40'),
    (45, '41 a 45'), (50, '46 a 50'), (55, '51 a 55'), (60, '56 a 60'),
]


def _rango_edad(fecha_nacimiento, hoy):
    if not fecha_nacimiento:
        return 'Sin dato'
    edad = hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
    for limite, etiqueta in RANGOS_EDAD:
        if edad <= limite:
            return etiqueta
    return 'Mayor a 60'


def _stats_discapacidad(request):
    servidores_visibles = filtrar_por_dependencia(
        ServidorPublico.objects.filter(activo=True), request.user, 'informacion_basica__dependencia_id'
    ).distinct()
    con_discapacidad = DiscapacidadServidor.objects.filter(
        servidor__in=servidores_visibles,
    ).exclude(discapacidad__clave__in=[0, 6]).select_related('servidor', 'discapacidad')

    servidor_ids = list(con_discapacidad.values_list('servidor_id', flat=True).distinct())
    total_con_discapacidad = len(servidor_ids)
    total_personas = servidores_visibles.count()

    hoy = date.today()
    fechas_nacimiento = ServidorPublico.objects.filter(pk__in=servidor_ids).values_list('pk', 'sexo', 'fecha_nacimiento')
    por_rango_map = {}
    for _pk, sexo, fecha_nac in fechas_nacimiento:
        rango = _rango_edad(fecha_nac, hoy)
        c = por_rango_map.setdefault(rango, {'hombres': 0, 'mujeres': 0, 'otro': 0})
        if sexo == 'MASCULINO':
            c['hombres'] += 1
        elif sexo == 'FEMENINO':
            c['mujeres'] += 1
        else:
            c['otro'] += 1
    orden_rangos = [e for _, e in RANGOS_EDAD] + ['Mayor a 60', 'Sin dato']
    por_rango_edad = []
    for r in orden_rangos:
        if r not in por_rango_map:
            continue
        c = por_rango_map[r]
        total_rango = sum(c.values())
        por_rango_edad.append({
            'label': r, **c, 'total': total_rango,
            'pct_hombres': round(c['hombres'] * 100 / total_rango, 1) if total_rango else 0,
            'pct_mujeres': round(c['mujeres'] * 100 / total_rango, 1) if total_rango else 0,
        })

    info_con_discapacidad = InformacionBasica.objects.filter(servidor_id__in=servidor_ids, activo=True)
    sueldo_por_sexo = [
        {
            'label': SEXO_LABELS.get(f['sexo'], f['sexo'] or 'Sin dato'),
            'promedio': round(f['promedio'], 2) if f['promedio'] is not None else 0,
        }
        for f in info_con_discapacidad.values('sexo').annotate(promedio=Avg('total_neto')).order_by('sexo')
    ]

    return {
        'total_personas': total_personas,
        'total_con_discapacidad': total_con_discapacidad,
        'porcentaje': round(total_con_discapacidad * 100 / total_personas, 1) if total_personas else 0,
        'por_tipo_sexo': _cruce_por_sexo(con_discapacidad, 'discapacidad__tipo', campo_sexo='servidor__sexo'),
        'por_rango_edad': por_rango_edad,
        'por_contratacion': _cruce_por_sexo(
            InformacionBasica.objects.filter(servidor_id__in=servidor_ids, activo=True), 'nombramiento__descripcion',
        ),
        'sueldo_por_sexo': sueldo_por_sexo,
    }


@login_required
@permiso_requerido('reporte_discapacidad')
def reporte_discapacidad(request):
    """Monitor de Discapacidad: servidores con discapacidad (excluye
    'Ninguna' y 'Sin respuesta' del catálogo) por tipo y sexo, por rango
    de edad, por tipo de contratación, y sueldo neto promedio por sexo —
    como el monitor de discapacidad del RUSP federal."""
    stats = _stats_discapacidad(request)
    return render(request, 'reportes/discapacidad.html', {'stats': stats, 'titulo': 'Monitor de Discapacidad'})


@login_required
@permiso_requerido('reporte_discapacidad')
def reporte_discapacidad_pdf(request):
    stats = _stats_discapacidad(request)
    pdf = generar_pdf_discapacidad(stats)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Monitor_Discapacidad_RESP.pdf"'
    return response


@login_required
def exportar_excel(request, tipo):
    wb = openpyxl.Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    if tipo == 'padron':
        ws.title = 'Padrón RESP'
        headers = ['Expediente', 'RFC', 'CURP', 'Nombre', 'Primer Apellido', 'Segundo Apellido',
                   'Fecha Nacimiento', 'Género', 'Correo Institucional', 'ISS', 'NSS', 'Activo']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        servidores_export = filtrar_por_dependencia(
            ServidorPublico.objects.filter(activo=True), request.user, 'informacion_basica__dependencia_id'
        ).distinct()
        for row, s in enumerate(servidores_export, 2):
            ws.cell(row=row, column=1, value=s.expediente)
            ws.cell(row=row, column=2, value=s.rfc)
            ws.cell(row=row, column=3, value=s.curp)
            ws.cell(row=row, column=4, value=s.nombre)
            ws.cell(row=row, column=5, value=s.primer_apellido)
            ws.cell(row=row, column=6, value=s.segundo_apellido or '')
            ws.cell(row=row, column=7, value=str(s.fecha_nacimiento))
            ws.cell(row=row, column=8, value=s.sexo)
            ws.cell(row=row, column=9, value=s.correo_institucional)
            ws.cell(row=row, column=10, value=s.iss)
            ws.cell(row=row, column=11, value=s.nss)
            ws.cell(row=row, column=12, value='Sí' if s.activo else 'No')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="RESP_{tipo}.xlsx"'
    wb.save(response)
    return response


# Override reporte_index to include items
def reporte_index(request):
    from django.contrib.auth.decorators import login_required
    items = [
        {'icono': '📄', 'nombre': 'Padrón General', 'desc': 'Listado completo de servidores activos', 'url': '/reportes/padron/'},
        {'icono': '⚖️', 'nombre': 'Declaración Patrimonial', 'desc': 'Sujetos obligados a declarar', 'url': '/reportes/declaracion-patrimonial/'},
        {'icono': '🤝', 'nombre': 'Entrega-Recepción', 'desc': 'Padrón de actas de entrega-recepción', 'url': '/reportes/entrega-recepcion/'},
        {'icono': '🚫', 'nombre': 'Bajas', 'desc': 'Registro histórico de bajas', 'url': '/reportes/bajas/'},
        {'icono': '🔄', 'nombre': 'Compatibilidad', 'desc': 'Servidores con más de una plaza', 'url': '/reportes/compatibilidad-horarios/'},
        {'icono': '📈', 'nombre': 'Estadísticas', 'desc': 'Indicadores generales del RESP', 'url': '/reportes/estadisticas/'},
        {'icono': '⚖️', 'nombre': 'Monitor de Paridad', 'desc': 'Distribución por género y sus cruces', 'url': '/reportes/paridad/'},
        {'icono': '📌', 'nombre': 'Ocupación de Plazas', 'desc': 'Ocupadas vs. vacantes, con porcentaje', 'url': '/reportes/ocupacion/'},
        {'icono': '♿', 'nombre': 'Monitor de Discapacidad', 'desc': 'Servidores con discapacidad por tipo, sexo y edad', 'url': '/reportes/discapacidad/'},
    ]
    return render(request, 'reportes/index.html', {'reporte_items': items, 'titulo': 'Reportes'})
